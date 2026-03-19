"""
Outbound Email Campaign Generator for Landscaper Lead Generation
Reads an Excel file with landscaper data, researches each website,
and generates personalized cold email sequences using Claude.
"""

import os
import sys
import time
import re
import traceback
from datetime import datetime
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import anthropic

# ─── CONFIG ────────────────────────────────────────────────────────────────────

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
MODEL = "claude-opus-4-6"

REQUEST_TIMEOUT = 12          # seconds for website fetch
REQUEST_DELAY   = 1.5         # seconds between website fetches
API_DELAY       = 2           # seconds between Claude API calls

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# ─── WEBSITE SCRAPER ───────────────────────────────────────────────────────────

def clean_text(text: str) -> str:
    """Collapse whitespace and strip."""
    return re.sub(r"\s+", " ", text).strip()


def fetch_website(url: str) -> dict:
    """
    Fetch a website and return structured content.
    Returns dict with keys: success, url, title, body_text, forms, error
    """
    result = {
        "success": False,
        "url": url,
        "title": "",
        "body_text": "",
        "has_contact_form": False,
        "has_booking": False,
        "has_cta": False,
        "page_meta": "",
        "error": "",
    }

    # Normalise URL
    if not url or url.strip() in ("", "N/A", "n/a", "-"):
        result["error"] = "No URL provided"
        return result

    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    result["url"] = url

    try:
        resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        resp.raise_for_status()
    except requests.exceptions.SSLError:
        try:
            resp = requests.get(url.replace("https://", "http://"), headers=HEADERS,
                                timeout=REQUEST_TIMEOUT, allow_redirects=True)
            resp.raise_for_status()
        except Exception as e:
            result["error"] = f"SSL + HTTP fallback failed: {str(e)[:120]}"
            return result
    except requests.exceptions.ConnectionError as e:
        result["error"] = f"Connection error: {str(e)[:120]}"
        return result
    except requests.exceptions.Timeout:
        result["error"] = "Request timed out"
        return result
    except requests.exceptions.HTTPError as e:
        result["error"] = f"HTTP {resp.status_code}: {str(e)[:80]}"
        return result
    except Exception as e:
        result["error"] = f"Unexpected: {str(e)[:120]}"
        return result

    try:
        soup = BeautifulSoup(resp.text, "lxml")

        # Remove noise
        for tag in soup(["script", "style", "nav", "footer", "head",
                          "noscript", "iframe", "svg"]):
            tag.decompose()

        # Title
        title_tag = soup.find("title")
        result["title"] = clean_text(title_tag.get_text()) if title_tag else ""

        # Meta description
        meta = soup.find("meta", attrs={"name": re.compile(r"description", re.I)})
        if meta and meta.get("content"):
            result["page_meta"] = clean_text(meta["content"])

        # Body text (limit to ~3000 chars to keep prompt manageable)
        body = soup.find("body")
        raw_text = body.get_text(separator=" ") if body else soup.get_text(separator=" ")
        result["body_text"] = clean_text(raw_text)[:3000]

        # Contact form detection
        forms = soup.find_all("form")
        form_texts = " ".join(str(f).lower() for f in forms)
        result["has_contact_form"] = bool(forms)

        # Booking signals
        booking_keywords = ["book", "schedule", "appointment", "calendar", "reserve",
                             "get a quote", "free quote", "request a quote", "estimate"]
        page_lower = resp.text.lower()
        result["has_booking"] = any(kw in page_lower for kw in booking_keywords)

        # CTA signals
        cta_keywords = ["call us", "contact us", "get started", "free estimate",
                        "call now", "get a quote", "request service"]
        result["has_cta"] = any(kw in page_lower for kw in cta_keywords)

        result["success"] = True

    except Exception as e:
        result["error"] = f"Parse error: {str(e)[:120]}"

    return result


# ─── CLAUDE EMAIL GENERATOR ────────────────────────────────────────────────────

def build_research_prompt(company: str, contact: str, site_data: dict) -> str:
    """Build the prompt asking Claude to analyse the site and write emails."""

    site_summary = ""
    if site_data["success"]:
        site_summary = f"""
WEBSITE DATA:
- Title: {site_data['title']}
- Meta description: {site_data['page_meta']}
- Has contact/lead form: {site_data['has_contact_form']}
- Has booking / quote CTA: {site_data['has_booking']}
- Has call-to-action: {site_data['has_cta']}
- Page text excerpt:
{site_data['body_text'][:2500]}
"""
    else:
        site_summary = f"WEBSITE FETCH FAILED: {site_data['error']}"

    return f"""You are an expert B2B cold email copywriter for a lead-filtering / lead-qualification service that helps landscaping companies stop wasting time on bad leads.

Company: {company}
Contact: {contact}
Website: {site_data['url']}

{site_summary}

YOUR TASK — produce a JSON object with exactly these keys:
1. "research_notes" — 2-4 sentences about what you observed on their site (services offered, lead capture quality, any obvious problems with how they handle enquiries). If the site failed to load, note that and describe what a typical landscaper site looks like.
2. "email_1" — the initial cold outreach email (100-150 words). Rules:
   - Subject line included at the top as "Subject: ..."
   - Reference something SPECIFIC from their website (a service, a gap, the booking flow, etc.)
   - Identify a lead capture / lead quality problem they likely have
   - Explain how we pre-qualify leads so they only speak to serious prospects ready to book
   - End with ONE clear CTA (e.g., "Worth a 15-min call this week?")
   - Tone: direct, peer-to-peer, no fluff
   - NO "I hope this email finds you well" or corporate opener
   - Sign off with: [Your Name] (placeholder)
3. "email_2" — follow-up 1 (send 3 days after email 1, ~80-100 words). Rules:
   - Subject: "Re: [original subject]" or a value-add subject
   - Add a specific insight: local landscaping market competition, seasonal lead spikes, or a stat about unqualified leads wasting field-crew time
   - Keep it punchy — one pain point, one nudge
   - No hard sell
4. "email_3" — follow-up 2 (send 5 days after email 2, ~80-100 words). Rules:
   - Lead with a brief case study or success metric (e.g., "One landscaper we work with cut their quote-to-close cycle by 40% after filtering...")
   - Soft close: "Happy to share how we'd do this for {company} specifically — just say the word."
   - This is the last touch; make it easy to respond

IMPORTANT:
- Return ONLY valid JSON, no markdown fences, no extra text.
- Make email_1 genuinely specific to this company/website. Generic = useless.
- If the website failed to load, write emails for a landscaper that probably has a basic site with a phone number and a "contact us" form — acknowledge you tried to find info on them.
"""


def generate_emails(client: anthropic.Anthropic, company: str, contact: str,
                    site_data: dict) -> dict:
    """Call Claude and return parsed JSON with research_notes + 3 emails."""
    prompt = build_research_prompt(company, contact, site_data)

    try:
        with client.messages.stream(
            model=MODEL,
            max_tokens=2048,
            thinking={"type": "adaptive"},
            messages=[{"role": "user", "content": prompt}],
        ) as stream:
            response = stream.get_final_message()

        raw = ""
        for block in response.content:
            if block.type == "text":
                raw += block.text

        # Strip markdown fences if present
        raw = re.sub(r"^```(?:json)?\s*", "", raw.strip())
        raw = re.sub(r"\s*```$", "", raw.strip())

        import json
        result = json.loads(raw)
        return result

    except Exception as e:
        return {
            "research_notes": f"ERROR generating content: {e}",
            "email_1": "ERROR — see research notes",
            "email_2": "ERROR — see research notes",
            "email_3": "ERROR — see research notes",
        }


# ─── EXCEL I/O ─────────────────────────────────────────────────────────────────

def read_input_file(path: str) -> list[dict]:
    """Read the input Excel file. Returns list of dicts."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Input file is empty.")

    # Try to detect header row
    header = [str(c).strip().lower() if c else "" for c in rows[0]]

    # Column name mapping (flexible)
    col_map = {}
    for i, h in enumerate(header):
        if any(k in h for k in ["company", "business", "name"]) and "contact" not in h:
            col_map.setdefault("company", i)
        elif any(k in h for k in ["contact", "first name", "person"]):
            col_map.setdefault("contact", i)
        elif any(k in h for k in ["website", "url", "web", "site"]):
            col_map.setdefault("website", i)
        elif any(k in h for k in ["phone", "tel", "mobile", "cell"]):
            col_map.setdefault("phone", i)

    records = []
    for row in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue  # skip blank rows
        record = {
            "Company Name": str(row[col_map.get("company", 0)] or "").strip(),
            "Contact Name": str(row[col_map.get("contact", 1)] or "").strip(),
            "Website URL":  str(row[col_map.get("website", 2)] or "").strip(),
            "Phone Number": str(row[col_map.get("phone", 3)] or "").strip(),
        }
        records.append(record)

    return records


def write_output_file(records: list[dict], output_path: str):
    """Write results to a formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Campaign"

    columns = [
        "Company Name", "Contact Name", "Website URL", "Phone Number",
        "Research Notes", "Email 1 (Initial)", "Email 2 (Follow-up Day 3)",
        "Email 3 (Follow-up Day 8)", "Status"
    ]

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    # Alternating row colours
    fill_even = PatternFill("solid", fgColor="D6E4F0")
    fill_odd  = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, rec in enumerate(records, 2):
        fill = fill_even if row_idx % 2 == 0 else fill_odd
        values = [
            rec.get("Company Name", ""),
            rec.get("Contact Name", ""),
            rec.get("Website URL", ""),
            rec.get("Phone Number", ""),
            rec.get("Research Notes", ""),
            rec.get("Email 1", ""),
            rec.get("Email 2", ""),
            rec.get("Email 3", ""),
            rec.get("Status", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    widths = [28, 20, 35, 18, 50, 70, 70, 70, 15]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # Row heights for content rows
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(records) + 2):
        ws.row_dimensions[row_idx].height = 120

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\n✅  Output saved → {output_path}")


# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  Landscaper Email Campaign Generator")
    print("=" * 60)

    # ── API Key ──────────────────────────────────────────────────
    if not ANTHROPIC_API_KEY:
        print("\n❌  ANTHROPIC_API_KEY environment variable not set.")
        print("    Set it with:  set ANTHROPIC_API_KEY=sk-ant-...")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    # ── Input file ───────────────────────────────────────────────
    if len(sys.argv) > 1:
        input_path = sys.argv[1]
    else:
        input_path = input("\nEnter path to input Excel file: ").strip().strip('"')

    if not os.path.isfile(input_path):
        print(f"❌  File not found: {input_path}")
        sys.exit(1)

    print(f"\n📂  Reading: {input_path}")
    records = read_input_file(input_path)
    print(f"✅  Loaded {len(records)} landscapers\n")

    # ── Output file ──────────────────────────────────────────────
    base, ext = os.path.splitext(input_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"{base}_CAMPAIGN_{timestamp}.xlsx"

    # ── Process each record ──────────────────────────────────────
    for i, rec in enumerate(records, 1):
        company = rec["Company Name"] or f"Landscaper #{i}"
        contact = rec["Contact Name"] or "there"
        url     = rec["Website URL"]

        print(f"[{i}/{len(records)}] {company}")
        print(f"  🌐  Fetching website: {url or '(none)'}")

        site_data = fetch_website(url)
        if site_data["success"]:
            print(f"  ✅  Site loaded — form:{site_data['has_contact_form']} "
                  f"booking:{site_data['has_booking']} cta:{site_data['has_cta']}")
        else:
            print(f"  ⚠️   Site fetch failed: {site_data['error']}")

        time.sleep(REQUEST_DELAY)

        print(f"  🤖  Generating emails with Claude...")
        output = generate_emails(client, company, contact, site_data)

        rec["Research Notes"] = output.get("research_notes", "")
        rec["Email 1"]        = output.get("email_1", "")
        rec["Email 2"]        = output.get("email_2", "")
        rec["Email 3"]        = output.get("email_3", "")
        rec["Status"]         = ""

        print(f"  ✅  Done\n")

        time.sleep(API_DELAY)

    # ── Save ─────────────────────────────────────────────────────
    write_output_file(records, output_path)
    print(f"\nAll {len(records)} landscapers processed.")
    print(f"Open your campaign file: {output_path}")


if __name__ == "__main__":
    main()

"""
DM Generator — Flask web app
Open the profile in Chrome → come back → click Generate.
Scrolls through the page, takes screenshots, sends to Groq Vision.
"""

import os, re, base64, random
from pathlib import Path
from flask import Flask, request, jsonify, render_template
from groq import Groq
from google import genai as _genai
from en_niche import (
    detect_niche as _en_detect_niche,
    detect_niche_from_name as _en_detect_niche_from_name,
    opener_matches_niche as _en_opener_matches_niche,
    NICHE_OPENER_INSTRUCTIONS as _EN_NICHE_OPENER_INSTRUCTIONS,
)
try:
    import pyperclip as _pyperclip
    _CLIPBOARD_OK = True
except ImportError:
    _CLIPBOARD_OK = False

# ── Load .env ──────────────────────────────────────────────────────────────────
_env = Path(__file__).parent / ".env"
if _env.exists():
    for line in _env.read_text().splitlines():
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

API_KEY          = os.environ.get("GROQ_API_KEY", "")
GEMINI_API_KEY   = os.environ.get("GEMINI_API_KEY", "")
CDP_PORT         = int(os.environ.get("CDP_PORT", "9222"))
DEFAULT_EXCEL    = os.environ.get("DEFAULT_EXCEL", "")
GRAMMAR_CHECK    = False  # set to True in dm_app_hu.py
GREETING_PREFIX  = "Hi"
STRICT_NAME_DETECTION = True    # mirror HU: only greet with a confidently detected name
ALWAYS_GREET     = True          # mirror HU: always greet ("Hi. " when no name found)
GREETING_NAME_FMT = "Hi, {name}. "   # ends with ". " → opener stays capitalised (mirror HU "Szia, {name}. ")
GREETING_SOLO_FMT = "Hi. "            # no-name fallback (mirror HU "Szia. ")
CUSTOM_NAME_EXTRACTOR = None               # if set, overrides extract_first_name (callable: text -> str)
CUSTOM_PROMPT_BUILDER = None               # if set, called as (full_prompt, text_data) -> str before vision call
CUSTOM_OPENER_POSTPROCESS = None           # if set, called as (opener, text_data) -> str after generation
VISION_MODELS  = ["meta-llama/llama-4-scout-17b-16e-instruct", "meta-llama/llama-4-maverick-17b-128e-instruct"]
TEXT_MODELS    = ["llama-3.3-70b-versatile", "llama-3.1-8b-instant"]
# Facts extraction is trivial bullet-pulling from explicit text — the fast 8B
# does it identically to the 70B, and it's the per-lead bottleneck. Quality of
# the OPENER comes from the vision call, which is untouched. 70B kept as fallback.
FACTS_MODELS   = ["llama-3.1-8b-instant", "llama-3.3-70b-versatile"]

# ── Message variant (A/B test toggle) ────────────────────────────────────────
# "full"     = current opener + full pitch body (default, unchanged behaviour).
# "stripped" = one-line direct variant. No personalised opener, so NO vision call
#              and NO screenshot — faster too. Reads:
#              "Hi [name]. I make AI quote chatbots for [trade] and I've already
#               built a sample for [company]. Can I send it over?"
# Persisted to settings.json so the toggle survives app restarts.
_SETTINGS_PATH = Path(__file__).parent / "settings.json"

def _load_variant() -> str:
    try:
        import json
        v = str(json.loads(_SETTINGS_PATH.read_text()).get("variant", "full")).strip().lower()
        return v if v in ("full", "stripped") else "full"
    except Exception:
        return "full"

def _save_variant(v: str) -> None:
    try:
        import json
        _SETTINGS_PATH.write_text(json.dumps({"variant": v}))
    except Exception as e:
        print(f"  Could not save variant setting: {e}")

MESSAGE_VARIANT = _load_variant()

PROMPT = """\
You write the opening line of a cold Facebook DM. Return ONLY the opening line — nothing else. No explanation, no emoji, no extra text.

You're looking at a screenshot of a UK trade business Facebook page. Look at the photo carefully. React to a finished job — like texting a mate who actually knows the trade.

PRIORITY — PROFESSIONAL DETAIL:
If you can see any of these, react to it — they're far more convincing than a generic "looks good":
- Brand name (see the list below)
- Material type (porcelain, Indian sandstone, EPDM, resin, composite, slate, lead, block paving, etc.)
- Placement / quality of the install (joint width, fall/slope, pipework run, edging course, pattern match)
- Size / scale of the project (how many panels, how big an area, a large system)
- A detail only a tradesperson would notice (ventilation, drainage fall, substructure, flashing detail, ridge line, gutter fall, lead soaker around a chimney, board spacing, membrane layer showing)

WINDOW / DOOR PHOTOS — look specifically for these:
- Frame colour: anthracite, dark brown, golden oak, white — how it suits the facade
- Opening type: tilt and turn, sliding / lift-and-slide, bi-fold — instantly professional
- Cill slope: is there a visible fall for drainage (a flat cill is a fault)
- Reveal depth: deeply set frame vs flush fit — only a tradesperson notices
- Frame-to-wall seal: taped or just rendered — taped is the premium detail
- On a front door: threshold meeting the floor finish, the seal compressing
- On a roof window (Velux/Fakro): flashing laid in, tiles returned around the frame
- On an integrated roller shutter: is the box built into the wall or sticking out

BRAND NAMES BY NICHE — if you see them, use them:
- Air con / split / heat pump: Daikin, Mitsubishi, LG, Samsung, Panasonic, Fujitsu, Midea, Gree, Toshiba
- Boiler / heating: Worcester Bosch, Vaillant, Ideal, Baxi, Viessmann, Glow-worm, Potterton
- Solar: LONGi, JA Solar, Jinko, Canadian Solar, Trina, SolarEdge (inverter), SMA, Fronius, GivEnergy
- Roofing: Marley, Redland, Sandtoft, Velux, Fakro, GRP/fibreglass, EPDM, lead
- Windows / doors: Schüco, Rehau, Veka, Liniar, Origin, Residence 9
- Paving / driveway: Marshalls, Bradstone, Brett, Tobermore, resin bound
- Decking: composite, Trex, Millboard, Cladco, hardwood

PROFESSIONAL EXAMPLES — this is what it looks like when someone knows the trade:
Air con / split (vary the focal point — the outdoor/condenser unit, the indoor head or cassette, the lagged line set, the wall bracket, the trunking, the condensate drain, the grille, where it's sited — NOT always "pipework run". Only what's actually visible):
"That Daikin condenser in Leeds came up mint, line set lagged dead neat."
"That Mitsubishi split in Bristol is sat lovely, indoor head level on the wall."
"That ceiling cassette in Leeds looks the part, grille sat flush in the ceiling."
"That outdoor unit in York is bang on, trunking clipped tight down the wall."
"That multi-split in Manchester looks proper tidy, two heads matched up neat."
"That condensate run in Sheffield is spot on, drain kept tidy off the unit."
Boiler / heating (vary the focal point — flue, condensate pipe, gas pipe, isolation valves, system/magnetic filter, the boiler model & position, cupboard mount, cylinder, manifold — NOT always "pipework on the wall". Only what's actually visible):
"That Worcester Bosch in Manchester looks proper tidy, flue routed clean through the wall."
"That combi swap in Leeds came up mint, isolation valves all lined up neat."
"That underfloor heating manifold in Leeds is bang on, neat loops off the ports."
"That Vaillant in Bristol looks the part, magnetic filter sat neat on the return."
"That boiler in York is sat lovely, condensate run kept tight to the wall."
"That cylinder install in Sheffield came up a treat, tidy tank and pump set."
"That gas run in Derby is proper tidy, clipped dead straight along the wall."
Solar:
"That LONGi array in York looks spot on, even pitch across the panels."
"That solar roof in Sheffield looks serious, proper symmetry on the layout."
Roofing:
"That Marley tiled roof in Leeds looks spot on, even coursing across the lot."
"That standing seam roof in York came up mint, clean verge line."
"That EPDM flat roof in Bolton looks proper tidy, neat edge trim."
"That lead flashing in Oxford is bang on, tight detail round the chimney."
Paving / driveway:
"That block paving in Coventry looks spot on, even joints right across."
"That resin driveway in Cardiff came up mint, proper clean finish."
Tiling / bathroom (vary the focal point — basin, suite, screen, niche, bath panel, tile pattern, brassware — instead of always grout. These show TONE and STRUCTURE only: only mention a detail you can ACTUALLY SEE, never borrow these specifics if they're not in the photo):
"That walk-in shower in Leeds came up mint, tiled niche sat dead level."
"That bathroom in Manchester is bang on, tidy floor-to-wall transition."
"That freestanding bath in Bristol looks the part, neat panel reveal round it."
"That metro tiling in York is proper tidy, crisp brick-bond lines."
"That wet room in Sheffield came up a treat, fall to the drain dead on."
"That vanity unit in Derby looks cracking, basin and brassware lined up neat."
"That herringbone floor in Leeds is bang on, tight joints right across."
Windows / doors:
"That Schüco anthracite window in Leeds looks spot on, deeply set frame."
"That Rehau tilt and turn in Bristol looks tidy, proper fall on the cill."
"That bi-fold in York came up mint, slim frame profile right across."
"That front door in Manchester looks bang on, tidy threshold to the floor."
"That Velux in Sheffield is spot on, flashing returned proper neat."
Garden / decking:
"That composite decking in Leeds looks spot on, even board spacing."
"That garden in York looks serious, proper line on the retaining levels."
Render / facade:
"That K Rend facade in Sheffield looks spot on, even coverage across the lot."
"That render in Manchester is proper tidy, sharp corner beads."
Painting / decorating:
"That exterior in Leeds came up spot on, even coverage right across."
"That interior in Manchester looks tidy, clean cutting-in round the frames."

TONE: Casual, British, working class. Personality. Like a text from someone who genuinely rates good work. Use words like: spot on, mint, cracking, proper, came up a treat, looks the part, bang on, tidy.

WHEN TO USE THE FALLBACK (compulsory, no exceptions):
- You can't see a finished trade job in the photo
- You only see text, a logo, a document, a certificate or an AI image
- The photo is a branded VAN / vehicle, a sign / banner, a business card, or a team / headshot photo — none of these are a finished job. Do NOT describe the van or the branding ("that silver car looks spot on" is wrong). → fallback
- You can't work out the location
- You CAN'T name a specific material, brand or professional detail — the location alone is NOT enough
- Generic words like "wallpaper", "painting", "tiling", "roof" ON THEIR OWN aren't enough — they need a colour, technique, brand or specific detail alongside. If there's none → fallback.
Fallback: "Came across your page in [location]." — town or city only.

STRICT RULES:
- Return ONLY the opening line. No other text.
- No emoji, no exclamation marks, no em dashes.
- Don't copy the examples word for word.
- 10-16 words max.
- EVERY sentence starts with "That" — the only exception is the fallback.
- BANNED headline style: "Nice job on the...", "Great work on the...", "Lovely finish on..." — banned.
- BANNED passive endings: "nicely finished", "well executed", "well laid", "nicely done" — banned. Just describe what you can actually see in a concrete second clause.
- VARY THE VERB: do NOT default to "looks spot on" every time — rotate naturally across "came up mint", "proper tidy", "bang on", "looks the part", "came up a treat", "is sat lovely", "spot on". Reusing the same phrase on every job reads like a template and kills it.
- NAME THE SPECIFIC THING you can see, and pick whatever is genuinely the STANDOUT in THIS photo. Do NOT let any one detail become the default across a batch — name the actual focal point of THIS particular image.
  - Bathroom menu (wide): the bath or freestanding tub, the shower or enclosure/screen, the WC, the flooring, the tiling layout (metro, herringbone, brick-bond), the vanity unit, the recessed niche, the towel rail/radiator, the mirror, the basin, the brassware/taps. OVERUSED, avoid as default: "grout lines", "basin / brassware".
  - Boiler / heating menu (wide): the flue, the condensate pipe, the gas pipe, the isolation valves, the system/magnetic filter, the boiler model & where it's sat, the cupboard mount, the cylinder, the manifold. OVERUSED, avoid as default: "pipework on the wall", "clean pipework".
  - Air con / HVAC menu (wide): the outdoor/condenser unit, the indoor head or cassette, the lagged line set, the wall bracket, the trunking/conduit, the condensate drain, the grille, where it's sited, the brand. OVERUSED, avoid as default: "tidy pipework run", "Daikin unit ... pipework run".
- BUT only describe what's genuinely visible — accuracy beats variety. If the only thing you can clearly see is plain tiling and grout, say that honestly, or use the fallback. NEVER invent a fixture, pattern, brand or colour just to sound specific or to avoid repeating yourself.
- BANNED generic endings: "looks well-finished", "looks well-built", "looks like a lot of work", "nice work", "good finish" — banned. The verb ALWAYS needs a concrete material, brand or professional term next to it.
- COMPULSORY: every opening line needs a COMMA-separated second clause with a concrete observation. e.g. "...came up mint, tidy detail round the frame." — the sentence never ends just on the location or the name of the object. The second clause must be a PHYSICAL detail of the work (a fit, a run, a level, a finish, a join) — NEVER a location or a phrase like "[town] installation" / "[town] job". If a location is all you've got, that's a fallback, not a second clause.
- No "roof job" — say the type: flat roof, pitched roof, lead roof, fibreglass roof, re-slate.
- No street names, road names, house numbers or postcodes — town or city only.
- No "in the photo" or "in the post". No "apparently" — never reference the caption. No commenting on their feed or post count.
- NEVER put the business / company name into the line. Describe the work, the brand, or the location — not the company's name. WRONG: "That DAJ Gas Services radiator...", "That Janie McK Heating Services installation...". RIGHT: "That radiator looks proper tidy, neat connections to the wall.", "That boiler swap came up mint, cylinder set sat neat." If you have no brand and no location, describe the work plainly or use the fallback.
- No inventing specs — only describe what you can ACTUALLY SEE in the photo. This rule OVERRIDES the "vary it" and "be specific" rules above: if being specific would mean guessing, don't — describe only what's really there, or use the fallback. A plain accurate line always beats a specific made-up one.
- For windows NEVER write "plastic window" — write "uPVC window" or the brand name.
- If you see a brand name in the photo, you MUST work it in — it's not optional.
- BRAND-NICHE MATCH: Air con / cooling brands (Daikin, Mitsubishi, Fujitsu, Midea, Gree, Toshiba) are only for air con businesses. Boiler / heating brands (Worcester Bosch, Vaillant, Baxi, Ideal, Viessmann) are only for heating businesses. Don't mix them.
- NEVER write "nice finish" — always name EXACTLY what's good: the joint width, the fit, the fall, the pipework run, the board spacing, the ridge line, etc.
- If a colour or material colour is visible (anthracite, graphite grey, natural, white, brown), name it too."""

# ── DM body copy — one per trade/sub-trade, no em dashes ─────────────────────

DM_BODY_ROOFING     = "heard a few roofers saying they lose jobs just cos someone else got back quicker. you find that?"
DM_BODY_DRIVEWAY    = "heard a few driveway lads saying enquiries come in then go dead after the quote. same with you?"
DM_BODY_PATIO       = "heard a few patio guys saying enquiries come in then go dead after the quote. same with you?"
DM_BODY_LANDSCAPERS = "heard a few landscapers saying they get enquiries then nothing after the quote. same with you?"
DM_BODY_BUILDERS    = "heard a few builders saying half their enquiries just want a ballpark then disappear. same with you?"
DM_BODY_RENOVATION  = "heard a few renovation lads saying they go price a job then just hear nothing back. same with you?"
DM_BODY_QUOTES      = "heard a few trades saying enquiries come in then go dead after the quote. same with you?"

# ── Follow-up sequences — 3 touches per trade ────────────────────────────────
# Touch 1 (Day 3):  new angle / peer curiosity — sounds like genuine interest
# Touch 2 (Day 10): social proof — concrete result from a similar business
# Touch 3 (Day 21): break-up / loss aversion — walk away, highest reply rate

FOLLOWUPS = {
    "roofers": [
        "Spoke to a roofer in Sheffield last week who said he loses half his enquiries just because he couldn't get back same day, does that happen much your end?",
        "Helped a roofing lad in Leeds go from missing calls to booking an extra 3-4 jobs a week, just by having enquiries handled while he was on the roof. Thought it might be relevant.",
        "Seems like the timing's not right. No worries, I'll leave it here. If you ever want to chat about getting more out of your enquiries, you know where I am.",
    ],
    "driveways": [
        "Spoke to a driveway firm in Birmingham who reckoned he was losing 2 jobs a week to lads who just quoted cheaper, do you find price shopping's got worse lately?",
        "Helped a driveway company in Manchester stop chasing dead quotes and start closing the ones that matter, 6 extra jobs booked in the first month. Might be worth a look.",
        "Looks like the timing's off. I'll leave this here. If it ever becomes relevant, feel free to drop me a message.",
    ],
    "patio": [
        "Spoke to a patio contractor in Bristol who said he spends half his time quoting jobs that go nowhere, do you get a lot of that round your way?",
        "Helped a landscaping firm in Leeds filter out the tyre-kickers and only spend time on serious enquiries, their close rate went from 1 in 5 to 1 in 2. Thought that might be useful.",
        "Looks like it's not the right time. No worries, I'll leave this here and you can always come back to it.",
    ],
    "landscapers": [
        "Spoke to a landscaper in Nottingham who said most of his best jobs still come through word of mouth but he can't make it consistent, is that the same for you?",
        "Helped a garden design firm in the Midlands turn their Facebook page into a steady stream of enquiries, went from sporadic word of mouth to 8-10 leads a week. Worth a chat?",
        "Seems like now isn't the time. I'll leave it there. If that changes, you know where I am.",
    ],
    "builders": [
        "Spoke to a builder in Leeds last week who said 3 out of 5 enquiries are people who just want a rough idea and disappear, do you find that's got worse recently?",
        "Helped a building firm in Manchester filter out the time-wasters and focus on serious enquiries, they went from 10 enquiries a week to 6 decent ones that actually go somewhere. Might be relevant.",
        "Looks like the timing's not right. No worries, I'll leave it here. Feel free to reach out if anything changes.",
    ],
    "renovation contractors": [
        "Spoke to a renovation contractor in London who said he spends days pricing jobs that go quiet after he sends the quote, does that happen much your end?",
        "Helped a renovation company in the South East cut their quote-to-job time in half by changing how they follow up, stopped losing jobs to silence. Thought it might be relevant.",
        "Seems like now isn't the right time. I'll leave it there. If you ever want to talk about it, feel free.",
    ],
    "quotes": [
        "Spoke to a contractor last week who said his quotes go cold about 60% of the time, no reply, no reason. Do you find that with yours?",
        "Helped a trades business in the Midlands go from chasing cold quotes to having clients chase them, changed one thing in how they follow up. Happy to share if it's useful.",
        "Looks like the timing's off. I'll leave this here. You know where I am if anything changes.",
    ],
}

# Follow-up sequence: 3 touches, widening gaps (≈ Day 3 / Day 10 / Day 21).
FOLLOWUP_TOUCHES = 3

def _clean_company_name(name: str) -> str:
    """Tidy a raw business name so it reads natural mid-sentence.
        'Pierce - Your Local Gas/Heating Engineer' -> 'Pierce'
        'Beagle Heating, Plumbing, Gas'            -> 'Beagle Heating'
        'Ava Bathrooms Ltd'                        -> 'Ava Bathrooms'
    Returns '' when nothing usable is left so the caller can fall back to
    'your business' instead of dumping a messy string into the message."""
    n = (name or "").strip()
    if not n:
        return ""
    # Cut at the first tagline separator ("Name - tagline", "Name | tagline"...)
    for sep in (" - ", " – ", " — ", " | ", " · ", " • ", ": ", " / "):
        i = n.find(sep)
        if i > 0:
            n = n[:i]
            break
    # Cut a comma list down to the lead name ("Beagle Heating, Plumbing, Gas")
    if "," in n:
        n = n.split(",", 1)[0]
    n = n.strip(" .-–—|/")
    # Drop a trailing parenthetical ("Air Con Services (UK)" -> "Air Con Services")
    n = re.sub(r"\s*\([^)]*\)\s*$", "", n).strip()
    # Drop trailing legal suffixes
    n = re.sub(r"\s*\b(ltd\.?|limited|llp|llc|plc|inc\.?|co\.?|company)\b\.?$", "", n, flags=re.I).strip()
    if not re.search(r"[A-Za-z]", n):
        return ""
    words = n.split()
    if len(words) > 6:           # still descriptor-laden — cap so it stays readable
        n = " ".join(words[:4])
    return n.strip()

def _followup_text(touch: int, company_name: str = "", niche: str = "") -> str:
    """Niche-specialised nudge — references the actual trade/job and the sample,
    in the same voice as the opener, so it reads tailored rather than generic.
    Two wordings per touch for spam-safety."""
    trade_noun, job, examples = _pitch_for(niche)
    who = _clean_company_name(company_name) or "your business"
    t = max(1, min(touch, FOLLOWUP_TOUCHES))
    if t == 1:
        return random.choice([
            f"Hi again, no rush at all. That sample I put together for {who} gives someone a rough price for {job} in about 30 seconds, then passes you their details. Happy to send it over whenever, just say the word.",
            f"Hey, no pressure on this. The sample I built for {who} shows someone a ballpark for {job} in seconds and sends you their info. Happy to drop it over whenever suits.",
        ])
    if t == 2:
        return random.choice([
            f"Just circling back once. It asks a couple of quick questions ({examples}), gives them a ballpark, and only passes you the serious ones. Want me to send the sample over?",
            f"Coming back to this one more time. The whole point is it sorts the time wasters from the real ones before they reach you, so you only quote the jobs worth quoting. Worth a quick look?",
        ])
    return random.choice([
        "All good if it's not for you, I'll leave it here. If you ever fancy a quick look at the sample, just give me a shout. Cheers.",
        "No worries if it's not your thing, I'll leave it here. The offer's open if you ever want a look. Cheers.",
    ])

VOICE_NOTE_PROMPT = """\
You write a short WhatsApp voice note script for someone reaching out cold to a UK landscaping business owner. A screenshot is being sent in the same message as this voice note. Output ONLY the script — no labels, no explanation, no quotes. Write it as natural spoken words.

This is the FIRST message. The recipient can already see the sender's business name on WhatsApp so there is no need to introduce yourself.

TONE: Casual, British, like a real person. Not a salesperson. Use natural British expressions. Sound like a mate who spotted something interesting, not an agency pitching a product.

STRUCTURE — 5 parts, under 90 words total (must be speakable in under 45 seconds):

1. GREETING + HOOK
Start with "Cheers [name]," — find the owner's first name from the About section ("Owner at", "Founder of", "Director"), the page name if it includes a person's name, or post captions. Never use the company name as a name.
Then reference ONE specific post — name the actual material, job type, and a visual detail or location from that post. Go as specific as possible.
GOOD: "just saw that resin bound driveway with the charcoal border you finished in Bristol, looks tidy" / "noticed that split-level garden with the porcelain and sleepers you posted"
BAD: "noticed you do landscaping in Devon" / "saw you work across Yorkshire" — these read the bio, not the posts.
End the hook naturally: "looks spot on" or similar. Keep it brief.

2. SCREENSHOT REFERENCE
Direct them to look at the screenshot. One sentence.
Example: "Have a quick look at that screenshot I sent — it is basically how an enquiry looks when it hits your phone."

3. WHAT THEY GET
Explain both sides in one sentence — what the customer gets and what the owner gets.
The customer gets a rough price instantly. The owner gets all their details — budget, postcode, job type — before picking up the phone.

4. WHAT IT DOES
One sentence. The point of it is to filter out tyre-kickers and save them grief.
Use natural British expressions: "tyre-kickers", "save you some grief", "takes the hassle out of it".
Do NOT use: leads, solution, system, software, AI, agent, scale, grow, qualified.

5. DISQUALIFY + CTA
"Might not be for you, but if you are interested in how it works, just reply and we will have a quick chat. Cheers."

RULES:
- Under 90 words total
- The hook must be specific to THIS page — not something generic you could send to any landscaper
- British vocabulary throughout — "mate", "cheers", "tyre-kickers", "the lot", "reckon", "grief" are all fair game
- No exclamation marks. No emojis. No over-the-top compliments.
- Sound like a real person, not a script

GOLD STANDARD EXAMPLE — match this tone exactly:
"Cheers Haydan. Just saw that garden transformation you posted over in West Yorkshire — looks spot on, mate. Have a quick look at that screenshot I sent — it is basically how an enquiry looks when it hits your phone. The customer gets a rough price instantly, and you get all their details — budget, postcode, the lot — before you even pick up the phone. I am building these for landscapers to help filter out the tyre-kickers and save you some grief. Might not be for you, but if you are interested in how it works, just reply and we will have a quick chat. Cheers."

Bad examples — never write these:
- "I have an AI solution that generates qualified leads" (jargon, wrong)
- "Would you be open to a quick chat?" (wrong CTA structure)
- "Saw you do landscaping in Manchester..." (no specific detail, too generic)
- "It asks questions and calculates a price..." (too specific, kills curiosity)"""

app = Flask(__name__)

# ── Capture existing Chrome tab ────────────────────────────────────────────────
def capture_profile_tab(hint_url: str = "") -> tuple:
    """
    Find the already-open Facebook/Instagram profile tab in Chrome.
    Scroll through it and take screenshots. Do NOT open a new tab.
    Returns (list of (bytes, mime_type), text_string)
    """
    import socket
    from playwright.sync_api import sync_playwright

    print("\nChecking Chrome debug port...")
    try:
        s = socket.create_connection(("127.0.0.1", CDP_PORT), timeout=3)
        s.close()
        print(f"Port {CDP_PORT} OPEN")
    except Exception:
        print(f"Port {CDP_PORT} CLOSED — Chrome is not running with debug port")
        return [], ""

    with sync_playwright() as p:
        try:
            browser = p.chromium.connect_over_cdp(f"http://localhost:{CDP_PORT}")
            context = browser.contexts[0]
            print(f"Connected. Open tabs: {len(context.pages)}")
            for pg in context.pages:
                print(f"  Tab: {pg.url[:80]}")
        except Exception as e:
            print(f"CDP connect error: {e}")
            return [], ""

        # Find the right tab — NEVER create a new one
        target_page = None

        # 1. Try to match hint URL if provided
        if hint_url:
            hint_clean = hint_url.rstrip("/").split("?")[0].lower()
            for pg in context.pages:
                if hint_clean in pg.url.lower():
                    target_page = pg
                    print(f"Matched hint URL: {pg.url}")
                    break

        # 2. Find any Facebook profile tab (not messenger/messages)
        if not target_page:
            for pg in context.pages:
                u = pg.url.lower()
                if "facebook.com/" in u and "messenger.com" not in u and "/messages" not in u:
                    target_page = pg
                    print(f"Found Facebook tab: {pg.url}")
                    break

        # 3. Find any Instagram tab (not DMs)
        if not target_page:
            for pg in context.pages:
                u = pg.url.lower()
                if "instagram.com/" in u and "/direct" not in u:
                    target_page = pg
                    print(f"Found Instagram tab: {pg.url}")
                    break

        if not target_page:
            print("No suitable Facebook/Instagram tab found")
            return [], ""

        try:
            target_page.bring_to_front()
            target_page.wait_for_timeout(600)

            # ── Extract text FIRST — fast, no scrolling needed ────────────────
            text = ""
            try:
                text = target_page.evaluate("""() => {
                    ['script','style','noscript','svg','iframe']
                        .forEach(t => document.querySelectorAll(t).forEach(el => el.remove()));
                    return document.body ? document.body.innerText : '';
                }""")
                text = re.sub(r"\n{3,}", "\n\n", text).strip()
                text = re.sub(r"(?im)^\s*[\d,\.]+\s*(followers?|likes?|reviews?|ratings?|check.?ins?|people follow|people like).*$", "", text)
                text = re.sub(r"\n{3,}", "\n\n", text).strip()[:12000]
                print(f"Text extracted: {len(text)} chars")
            except Exception as e:
                print(f"Text extraction skipped: {e}")

            # ── Always screenshot the first post area — vision sees what text misses ──
            screenshots = []
            try:
                # Scroll to where the first post typically lives on a Facebook Page
                target_page.evaluate("window.scrollTo(0, 850)")
                target_page.wait_for_timeout(1400)   # let images render
                shot = target_page.screenshot(type="jpeg", quality=72)
                screenshots.append((shot, "image/jpeg"))
                print(f"First-post screenshot: {len(shot)} bytes")
                target_page.evaluate("window.scrollTo(0, 0)")
            except Exception as _se:
                print(f"Screenshot failed: {_se}")

            return screenshots, text

        except Exception as e:
            print(f"Capture error: {e}")
            return [], ""

# ── Name extractor ────────────────────────────────────────────────────────────
# Common UK / English first names. A detected name is only used if it appears
# here — so we greet "Hi, <name>" only when a real first name is genuinely in
# the company name or bio, and never guess a non-name word ("Designer",
# "Newcastle", "Bathroom", etc.). Lowercase for case-insensitive matching.
FIRST_NAMES = {
    # ── male ──────────────────────────────────────────────────────────────────
    "aaron","adam","adrian","aidan","alan","albert","alex","alexander","alfie","alfred",
    "ali","allan","andrew","andy","angus","anthony","antony","archie","arthur","ashley",
    "barry","ben","benjamin","bernard","bill","billy","bobby","brad","bradley","brandon",
    "brendan","brett","brian","bruce","bryan","callum","cameron","carl","charles","charlie",
    "chris","christian","christopher","clive","colin","connor","conor","craig","curtis","dale",
    "damian","damien","daniel","danny","darren","darryl","dave","david","dean","declan",
    "dennis","derek","dom","dominic","donald","douglas","duncan","dylan","eddie","edward",
    "elliot","elliott","eric","ethan","ewan","felix","finlay","finn","francis","frank",
    "frankie","fred","freddie","frederick","gareth","garry","gary","gavin","geoff","geoffrey",
    "george","gerald","gerard","glen","glenn","gordon","graeme","graham","grant","greg",
    "gregory","guy","harold","harry","harvey","henry","howard","hugh","hugo","ian",
    "isaac","ivan","jack","jacob","jake","james","jamie","jared","jason","jay",
    "jeff","jeffrey","jeremy","jim","jimmy","joe","joel","john","johnny","jon",
    "jonathan","jonny","jordan","joseph","josh","joshua","julian","justin","karl","keith",
    "ken","kenneth","kevin","kieran","kyle","lance","larry","lee","leon","leonard",
    "leslie","lewis","liam","lloyd","logan","louis","luca","lucas","luke","malcolm",
    "marc","marcus","mark","martin","mason","matt","matthew","maurice","max","maxwell",
    "micah","michael","mick","mike","miles","mitchell","morgan","nathan","nathaniel","neil",
    "nick","nicholas","nigel","noah","norman","oliver","ollie","oscar","owen","pablo",
    "patrick","paul","percy","perry","pete","peter","philip","phillip","ralph","ray",
    "raymond","reece","reg","reginald","rhys","richard","rick","ricky","rob","robbie",
    "robert","robin","rod","rodney","roger","ronald","ronnie","rory","ross","roy",
    "russell","ryan","sam","samuel","scott","sean","seb","sebastian","shane","shaun",
    "sidney","simon","stanley","stephen","steve","steven","stewart","stuart","ted","terence",
    "terry","theo","thomas","tim","timothy","toby","todd","tom","tommy","tony",
    "travis","trevor","tristan","troy","tyler","vernon","victor","vincent","wade","walter",
    "warren","wayne","wesley","will","william","willie","zac","zach",
    # ── female ────────────────────────────────────────────────────────────────
    "abbie","abby","abigail","adele","aimee","alana","alexandra","alexis","alice","alicia",
    "alison","amanda","amber","amelia","amy","andrea","angela","angie","anita","ann",
    "anna","anne","annette","annie","april","audrey","barbara","becky","bella","beth",
    "bethan","bethany","betty","beverley","bonnie","brenda","bridget","brooke","carla","carly",
    "carol","caroline","carolyn","carrie","casey","catherine","cathy","charlene","charlotte","chelsea",
    "cheryl","chloe","christina","christine","claire","clara","clare","colleen","connie","courtney",
    "daisy","dana","danielle","dawn","deborah","debbie","debra","denise","diana","diane",
    "donna","doreen","doris","dorothy","ebony","eden","eileen","elaine","eleanor","elena",
    "eliza","elizabeth","ella","ellen","ellie","eloise","elsie","emily","emma","erica",
    "erin","esme","esther","eva","eve","evelyn","faith","faye","felicity","fern",
    "fiona","florence","frances","francesca","freya","gabrielle","gail","gemma","georgia","georgina",
    "gillian","gina","gloria","grace","gwen","hannah","harriet","hayley","hazel","heather",
    "heidi","helen","helena","holly","hope","imogen","irene","iris","isabel","isabella",
    "isabelle","isla","ivy","jacqueline","jade","jane","janet","janice","jasmine","jean",
    "jenna","jennifer","jenny","jessica","jill","joan","joanna","joanne","jodie","josephine",
    "joy","joyce","judith","judy","julia","julie","june","karen","kate","katherine",
    "kathleen","kathryn","kathy","katie","katrina","kay","kayleigh","keeley","kelly","kerry",
    "kim","kimberley","kirsten","kirsty","lacey","lara","laura","lauren","leah","leanne",
    "lesley","libby","lillian","lily","linda","lindsay","lindsey","lisa","lois","lola",
    "lorna","lorraine","louise","lucy","lydia","lynda","lynn","lynne","mabel","madeleine",
    "maggie","mandy","margaret","maria","marie","marilyn","marion","marjorie","marlene","martha",
    "mary","matilda","maureen","maya","megan","melanie","melissa","mia","michelle","millie",
    "miranda","miriam","molly","monica","nadia","nancy","naomi","natalie","natasha","nicola",
    "nicole","nina","nora","norma","olive","olivia","paige","pamela","patricia","paula",
    "pauline","pearl","peggy","penelope","penny","phoebe","phyllis","polly","poppy","rachael",
    "rachel","rebecca","renee","rhona","rita","roberta","robyn","rosa","rose","rosemary",
    "rosie","roxanne","ruby","ruth","sabrina","sadie","sally","samantha","sandra","sara",
    "sarah","sasha","scarlett","shannon","sharon","sheila","shelley","shirley","sian","sienna",
    "simone","sofia","sonia","sophia","sophie","stacey","stella","stephanie","sue","susan",
    "suzanne","sylvia","tamara","tania","tanya","tara","teresa","tessa","theresa","tiffany",
    "tina","tracey","tracy","valerie","vanessa","vera","veronica","vicky","victoria","violet",
    "virginia","vivienne","wendy","whitney","willow","yvonne","zara","zoe",
}


def _name_from_company(company_name: str) -> str:
    """A person's first name inside a company name, or ''.

    Whitelist-only (FIRST_NAMES): we use a token from the company name as the
    greeting name only if it's a known first name. Structural 'Firstname
    Surname' guessing was tried and removed — it mistook brand words like
    'Plumbergy', 'Westcountry' and 'Queen' for names. If a real name is missed,
    add it to FIRST_NAMES.
    """
    for t in re.findall(r"[A-Za-z]+", company_name or ""):
        if t.lower() in FIRST_NAMES:
            return t.capitalize()
    return ""


def extract_first_name(text: str, company_name: str = "") -> str:
    """
    Return the owner's first name ONLY if it's genuinely present in the company
    name or the bio. Never guesses a trade/place/generic word. Returns "" when
    no real first name is found → caller falls back to a plain "Hi.".
    """
    if CUSTOM_NAME_EXTRACTOR:
        return CUSTOM_NAME_EXTRACTOR(text)

    # 1) A first name inside the COMPANY NAME (whitelist, then Firstname-Surname)
    name = _name_from_company(company_name)
    if name:
        return name

    # 2) An explicit owner line / self-intro in the BIO — captured name must be
    #    a known first name, otherwise we ignore it (no guessing).
    bio_patterns = [
        r"\b(?:Owner|Founder|Co[\s\-]?founder|Director|Proprietor|Manager|Run by|MD)\b"
        r"[^\n]{0,15}?\b([A-Z][a-z]{1,14})\b",
        r"\b(?:I['’]?m|I am|my name is|name['’]s)\s+([A-Z][a-z]{1,14})\b",
    ]
    for pat in bio_patterns:
        m = re.search(pat, text or "")
        if m and m.group(1).lower() in FIRST_NAMES:
            return m.group(1).capitalize()

    return ""

# ── Location extractor ───────────────────────────────────────────────────────
UK_PLACES = [
    # Major cities
    "London","Manchester","Birmingham","Leeds","Liverpool","Sheffield","Bristol",
    "Edinburgh","Glasgow","Cardiff","Belfast","Newcastle","Nottingham","Leicester",
    "Coventry","Bradford","Stoke","Wolverhampton","Derby","Southampton","Portsmouth",
    "Brighton","Oxford","Cambridge","Norwich","Exeter","Plymouth","Swindon",
    "Gloucester","Cheltenham","Bournemouth","Peterborough","Luton","Milton Keynes",
    "Reading","Slough","Ipswich","Middlesbrough","Sunderland","Bolton","Blackpool",
    "Blackburn","Oldham","Rochdale","Stockport","Huddersfield","York","Lincoln",
    "Swansea","Newport","Dundee","Aberdeen","Inverness","Stirling","Perth",
    # Counties / regions
    "Devon","Cornwall","Somerset","Dorset","Wiltshire","Hampshire","Kent","Surrey",
    "Sussex","Essex","Suffolk","Norfolk","Hertfordshire","Bedfordshire","Berkshire",
    "Buckinghamshire","Oxfordshire","Gloucestershire","Worcestershire","Warwickshire",
    "Northamptonshire","Cambridgeshire","Leicestershire","Nottinghamshire","Derbyshire",
    "Staffordshire","Shropshire","Cheshire","Lancashire","Yorkshire","Cumbria",
    "Durham","Northumberland","Herefordshire","Lincolnshire","Rutland",
    "East Yorkshire","West Yorkshire","South Yorkshire","North Yorkshire",
    "West Midlands","East Midlands","East Anglia",
    # Common towns
    "Guildford","Woking","Crawley","Horsham","Tunbridge Wells","Maidstone","Canterbury",
    "Rochester","Basildon","Chelmsford","Colchester","Southend","St Albans","Watford",
    "Hemel Hempstead","Stevenage","Harlow","Aylesbury","High Wycombe","Bracknell",
    "Basingstoke","Winchester","Eastbourne","Worthing","Hastings","Folkestone",
    "Bath","Taunton","Yeovil","Poole","Weymouth","Salisbury","Chippenham","Swindon",
    "Torquay","Paignton","Barnstaple","Truro","Newquay","Penzance","Falmouth",
    "Hereford","Shrewsbury","Telford","Worcester","Redditch","Kidderminster",
    "Leamington Spa","Rugby","Nuneaton","Tamworth","Lichfield","Burton upon Trent",
    "Stafford","Walsall","West Bromwich","Dudley","Solihull","Sutton Coldfield",
    "Crewe","Chester","Warrington","Wigan","Preston","Burnley","Lancaster","Carlisle",
    "Harrogate","Wakefield","Barnsley","Doncaster","Rotherham","Chesterfield",
    "Mansfield","Newark","Grimsby","Scunthorpe","Hull","Scarborough","Whitby",
    "Darlington","Durham","Stockton","Gateshead","South Shields","Hartlepool",
    "Berwick","Hexham","Morpeth",
]

UK_NEIGHBOURS = {
    # Counties
    "Essex":             ["Suffolk", "Kent", "Hertfordshire", "Cambridgeshire"],
    "Kent":              ["Surrey", "East Sussex", "West Sussex", "Essex"],
    "Surrey":            ["Kent", "East Sussex", "West Sussex", "Hampshire", "Berkshire"],
    "East Sussex":       ["Kent", "Surrey", "West Sussex"],
    "West Sussex":       ["Hampshire", "Surrey", "East Sussex"],
    "Sussex":            ["Kent", "Surrey", "Hampshire"],
    "Devon":             ["Cornwall", "Somerset", "Dorset"],
    "Cornwall":          ["Devon", "Somerset"],
    "Somerset":          ["Devon", "Dorset", "Wiltshire", "Gloucestershire"],
    "Dorset":            ["Somerset", "Wiltshire", "Hampshire", "Devon"],
    "Hampshire":         ["Dorset", "Wiltshire", "Berkshire", "Surrey", "West Sussex"],
    "Wiltshire":         ["Somerset", "Dorset", "Hampshire", "Berkshire", "Gloucestershire", "Oxfordshire"],
    "Berkshire":         ["Wiltshire", "Hampshire", "Surrey", "Buckinghamshire", "Oxfordshire"],
    "Oxfordshire":       ["Berkshire", "Gloucestershire", "Warwickshire", "Northamptonshire", "Buckinghamshire"],
    "Buckinghamshire":   ["Oxfordshire", "Berkshire", "Hertfordshire", "Northamptonshire"],
    "Hertfordshire":     ["Essex", "Buckinghamshire", "Bedfordshire", "Cambridgeshire"],
    "Bedfordshire":      ["Hertfordshire", "Buckinghamshire", "Northamptonshire", "Cambridgeshire"],
    "Cambridgeshire":    ["Hertfordshire", "Bedfordshire", "Northamptonshire", "Lincolnshire", "Norfolk", "Suffolk"],
    "Suffolk":           ["Norfolk", "Cambridgeshire", "Essex"],
    "Norfolk":           ["Suffolk", "Cambridgeshire", "Lincolnshire"],
    "Lincolnshire":      ["Norfolk", "Cambridgeshire", "Leicestershire", "Nottinghamshire", "Yorkshire"],
    "Northamptonshire":  ["Oxfordshire", "Buckinghamshire", "Bedfordshire", "Cambridgeshire", "Leicestershire", "Warwickshire"],
    "Leicestershire":    ["Northamptonshire", "Lincolnshire", "Nottinghamshire", "Derbyshire", "Warwickshire"],
    "Nottinghamshire":   ["Lincolnshire", "Leicestershire", "Derbyshire", "Yorkshire"],
    "Derbyshire":        ["Nottinghamshire", "Leicestershire", "Staffordshire", "Cheshire", "Yorkshire"],
    "Staffordshire":     ["Derbyshire", "Leicestershire", "Warwickshire", "Worcestershire", "Shropshire", "Cheshire"],
    "Warwickshire":      ["Northamptonshire", "Leicestershire", "Staffordshire", "Worcestershire", "Oxfordshire", "Gloucestershire"],
    "Worcestershire":    ["Warwickshire", "Staffordshire", "Shropshire", "Herefordshire", "Gloucestershire"],
    "Herefordshire":     ["Worcestershire", "Shropshire", "Gloucestershire"],
    "Shropshire":        ["Staffordshire", "Worcestershire", "Herefordshire", "Cheshire"],
    "Gloucestershire":   ["Herefordshire", "Worcestershire", "Warwickshire", "Oxfordshire", "Wiltshire", "Somerset"],
    "Cheshire":          ["Staffordshire", "Derbyshire", "Lancashire", "Shropshire"],
    "Lancashire":        ["Cheshire", "Yorkshire", "Cumbria"],
    "North Yorkshire":   ["Lancashire", "Durham", "Cumbria", "East Yorkshire", "West Yorkshire"],
    "West Yorkshire":    ["Lancashire", "Derbyshire", "South Yorkshire", "North Yorkshire"],
    "South Yorkshire":   ["Nottinghamshire", "Derbyshire", "West Yorkshire", "Lincolnshire"],
    "East Yorkshire":    ["Lincolnshire", "North Yorkshire", "South Yorkshire"],
    "Yorkshire":         ["Lancashire", "Lincolnshire", "Nottinghamshire", "Derbyshire", "Durham"],
    "Cumbria":           ["Lancashire", "Yorkshire", "Durham", "Northumberland"],
    "Durham":            ["Cumbria", "Yorkshire", "Northumberland"],
    "Northumberland":    ["Durham", "Cumbria"],
    # Cities / large towns
    "London":            ["Essex", "Kent", "Surrey", "Hertfordshire"],
    "Manchester":        ["Cheshire", "Lancashire", "Derbyshire"],
    "Birmingham":        ["Warwickshire", "Staffordshire", "Worcestershire"],
    "Leeds":             ["Lancashire", "Harrogate", "Bradford", "Wakefield"],
    "Liverpool":         ["Lancashire", "Cheshire"],
    "Sheffield":         ["Derbyshire", "Nottinghamshire", "Barnsley"],
    "Bristol":           ["Somerset", "Gloucestershire", "Wiltshire"],
    "Newcastle":         ["Durham", "Northumberland", "Sunderland"],
    "Sunderland":        ["Durham", "Northumberland", "Newcastle"],
    "Norwich":           ["Norfolk", "Suffolk", "Cambridgeshire"],
    "Exeter":            ["Somerset", "Cornwall", "Dorset"],
    "Plymouth":          ["Cornwall", "Somerset"],
    "Brighton":          ["East Sussex", "West Sussex", "Surrey"],
    "Southampton":       ["Hampshire", "Dorset", "Wiltshire"],
    "Portsmouth":        ["Hampshire", "West Sussex", "Surrey"],
    "York":              ["North Yorkshire", "East Yorkshire", "West Yorkshire"],
    "Hull":              ["East Yorkshire", "Lincolnshire"],
    "Nottingham":        ["Derbyshire", "Leicestershire", "Lincolnshire"],
    "Leicester":         ["Northamptonshire", "Warwickshire", "Lincolnshire"],
    "Derby":             ["Nottinghamshire", "Staffordshire", "Leicestershire"],
    "Oxford":            ["Berkshire", "Buckinghamshire", "Gloucestershire"],
    "Cambridge":         ["Suffolk", "Hertfordshire", "Bedfordshire"],
    "Coventry":          ["Warwickshire", "Northamptonshire", "Staffordshire"],
    "Stoke":             ["Cheshire", "Derbyshire", "Shropshire"],
    "Wolverhampton":     ["Staffordshire", "Worcestershire", "Shropshire"],
    "Reading":           ["Berkshire", "Oxfordshire", "Hampshire"],
    "Gloucester":        ["Worcestershire", "Herefordshire", "Wiltshire"],
    "Cheltenham":        ["Worcestershire", "Oxfordshire", "Wiltshire"],
    "Bournemouth":       ["Dorset", "Somerset", "Hampshire"],
    "Poole":             ["Dorset", "Somerset", "Hampshire"],
    "Bath":              ["Wiltshire", "Gloucestershire", "Somerset"],
    "Taunton":           ["Devon", "Dorset", "Wiltshire"],
    "Torquay":           ["Cornwall", "Somerset"],
    "Barnstaple":        ["Cornwall", "Somerset"],
    "Preston":           ["Lancashire", "Cheshire", "Yorkshire"],
    "Blackpool":         ["Lancashire", "Cumbria"],
    "Blackburn":         ["Lancashire", "Yorkshire", "Cheshire"],
    "Bolton":            ["Lancashire", "Cheshire"],
    "Chester":           ["Cheshire", "Shropshire", "Lancashire"],
    "Warrington":        ["Cheshire", "Lancashire"],
    "Shrewsbury":        ["Staffordshire", "Cheshire", "Herefordshire"],
    "Telford":           ["Staffordshire", "Worcestershire", "Cheshire"],
    "Worcester":         ["Herefordshire", "Warwickshire", "Gloucestershire"],
    "Hereford":          ["Worcestershire", "Shropshire", "Gloucestershire"],
    "Harrogate":         ["West Yorkshire", "Lancashire", "Durham"],
    "Wakefield":         ["West Yorkshire", "South Yorkshire", "Lancashire"],
    "Huddersfield":      ["West Yorkshire", "Lancashire", "Derbyshire"],
    "Doncaster":         ["Lincolnshire", "Nottinghamshire", "West Yorkshire"],
    "Barnsley":          ["West Yorkshire", "Derbyshire", "Nottinghamshire"],
    "Scarborough":       ["North Yorkshire", "East Yorkshire"],
    "Middlesbrough":     ["Durham", "North Yorkshire"],
    "Darlington":        ["Durham", "North Yorkshire", "Cumbria"],
    "Gateshead":         ["Durham", "Northumberland"],
    "Carlisle":          ["Cumbria", "Northumberland", "Durham"],
    "Colchester":        ["Suffolk", "Hertfordshire", "Cambridgeshire"],
    "Chelmsford":        ["Suffolk", "Hertfordshire", "Kent"],
    "Ipswich":           ["Norfolk", "Cambridgeshire", "Essex"],
    "Guildford":         ["Kent", "Hampshire", "Berkshire"],
    "Maidstone":         ["Surrey", "East Sussex", "Essex"],
    "Canterbury":        ["East Sussex", "Surrey", "Essex"],
    "Hastings":          ["Kent", "Surrey"],
    "Eastbourne":        ["Kent", "Surrey", "West Sussex"],
    "Tunbridge Wells":   ["East Sussex", "Surrey", "Essex"],
    "Truro":             ["Devon", "Somerset"],
    "Newquay":           ["Devon", "Somerset"],
    "Penzance":          ["Devon"],
    "Falmouth":          ["Devon", "Somerset"],
    "Basingstoke":       ["Surrey", "Berkshire", "Wiltshire"],
    "Winchester":        ["Hampshire", "Berkshire", "Wiltshire"],
    # Scotland
    "Edinburgh":         ["Glasgow", "Stirling", "Fife", "Livingston"],
    "Glasgow":           ["Edinburgh", "Stirling", "Paisley", "Hamilton"],
    "Dundee":            ["Perth", "Fife", "St Andrews", "Angus"],
    "Aberdeen":          ["Dundee", "Inverness", "Elgin"],
    "Inverness":         ["Aberdeen", "Elgin", "Aviemore"],
    "Stirling":          ["Edinburgh", "Glasgow", "Perth", "Falkirk"],
    "Perth":             ["Dundee", "Stirling", "Fife", "Crieff"],
    "Falkirk":           ["Edinburgh", "Glasgow", "Stirling"],
    "Livingston":        ["Edinburgh", "Falkirk", "Stirling"],
    "Paisley":           ["Glasgow", "Ayr", "Kilmarnock"],
    "Hamilton":          ["Glasgow", "Motherwell", "Lanark"],
    "Fife":              ["Edinburgh", "Dundee", "Perth", "St Andrews"],
    "St Andrews":        ["Dundee", "Fife", "Perth"],
    "Ayr":               ["Glasgow", "Paisley", "Kilmarnock"],
    "Kilmarnock":        ["Ayr", "Glasgow", "Paisley"],
    "Scotland":          ["Glasgow", "Stirling", "Perth", "Fife"],
    # Wales
    "Cardiff":           ["Newport", "Swansea", "Bristol"],
    "Swansea":           ["Cardiff", "Newport", "Carmarthen"],
    "Newport":           ["Cardiff", "Bristol", "Herefordshire"],
    "Wrexham":           ["Chester", "Shrewsbury", "Flintshire"],
    "Bangor":            ["Caernarfon", "Conwy", "Anglesey"],
    "Carmarthen":        ["Swansea", "Pembroke", "Ceredigion"],
    "Wales":             ["Cardiff", "Swansea", "Newport"],
    # Northern Ireland
    "Belfast":           ["Lisburn", "Bangor", "Newtownabbey"],
    "Derry":             ["Coleraine", "Limavady", "Strabane"],
    "Northern Ireland":  ["Belfast", "Lisburn", "Bangor"],
}

# Map specific towns to their county for neighbour lookup
TOWN_TO_COUNTY = {
    "Poulton le Fylde": "Lancashire", "Poulton-le-Fylde": "Lancashire",
    "Lytham": "Lancashire", "Lytham St Annes": "Lancashire",
    "Fleetwood": "Lancashire", "Cleveleys": "Lancashire",
    "Southend": "Essex", "Basildon": "Essex", "Harlow": "Essex",
    "Woking": "Surrey", "Crawley": "West Sussex", "Horsham": "West Sussex",
    "Worthing": "West Sussex",
    "Rochester": "Kent", "Folkestone": "Kent",
    "Paignton": "Devon", "Bideford": "Devon",
    "Yeovil": "Somerset",
    "Weymouth": "Dorset",
    "Salisbury": "Wiltshire", "Chippenham": "Wiltshire",
    "Aylesbury": "Buckinghamshire", "High Wycombe": "Buckinghamshire",
    "Bracknell": "Berkshire", "Slough": "Berkshire",
    "Luton": "Bedfordshire", "St Albans": "Hertfordshire",
    "Watford": "Hertfordshire", "Hemel Hempstead": "Hertfordshire",
    "Stevenage": "Hertfordshire",
    "Peterborough": "Cambridgeshire",
    "Grimsby": "Lincolnshire", "Scunthorpe": "Lincolnshire", "Lincoln": "Lincolnshire",
    "Wigan": "Lancashire", "Burnley": "Lancashire", "Lancaster": "Lancashire",
    "Crewe": "Cheshire", "Stockport": "Cheshire",
    "Oldham": "Lancashire", "Rochdale": "Lancashire",
    "Bradford": "West Yorkshire", "Halifax": "West Yorkshire",
    "Rotherham": "South Yorkshire", "Chesterfield": "Derbyshire",
    "Mansfield": "Nottinghamshire", "Newark": "Nottinghamshire",
    "Tamworth": "Staffordshire", "Stafford": "Staffordshire",
    "Lichfield": "Staffordshire", "Burton upon Trent": "Staffordshire",
    "Walsall": "Staffordshire", "West Bromwich": "West Midlands",
    "Dudley": "West Midlands", "Solihull": "Warwickshire",
    "Sutton Coldfield": "Warwickshire",
    "Redditch": "Worcestershire", "Kidderminster": "Worcestershire",
    "Leamington Spa": "Warwickshire", "Rugby": "Warwickshire", "Nuneaton": "Warwickshire",
    "Hexham": "Northumberland", "Berwick": "Northumberland", "Morpeth": "Northumberland",
    "Hartlepool": "Durham", "Stockton": "Durham", "South Shields": "Durham",
    "Whitby": "North Yorkshire",
    "Milton Keynes": "Buckinghamshire",
    "Swindon": "Wiltshire",
    "Gloucester": "Gloucestershire",
    "Swansea": "Swansea", "Newport": "Newport", "Cardiff": "Cardiff",
    "Dundee": "Dundee", "Aberdeen": "Aberdeen", "Inverness": "Inverness",
    "Stirling": "Stirling", "Perth": "Perth", "Edinburgh": "Edinburgh",
    "Glasgow": "Glasgow", "Belfast": "Belfast",
    "Falkirk": "Falkirk", "Livingston": "Livingston", "Paisley": "Paisley",
    "Hamilton": "Hamilton", "Ayr": "Ayr", "Kilmarnock": "Kilmarnock",
    "St Andrews": "St Andrews", "Wrexham": "Wrexham", "Carmarthen": "Carmarthen",
}


def extract_prospect_location(opener: str, text: str) -> str:
    """
    Extract the prospect's own location from the opener or page text.
    """
    # 1. From the opener — "in Bristol", "up in Leeds", "over in Devon"
    m = re.search(
        r'\b(?:in|up in|down in|over in)\s+([A-Z][a-zA-Z ]{2,30}?)(?:\s*[,.]|\s*$|\s+last\b|\s+the\s+other)',
        opener
    )
    if m:
        loc = m.group(1).strip().rstrip(",. ")
        words = loc.split()
        if 1 <= len(words) <= 3 and words[0][0].isupper():
            return loc

    # 2. Page text — "based in X", "serving X", "covering X"
    m = re.search(
        r'(?:based in|serving|located in|covering|working in|work across|operating in)'
        r'\s+([A-Z][a-zA-Z ,&]{2,40}?)(?:\s*[\n,]|\.|\band\b)',
        text, re.I
    )
    if m:
        loc = m.group(1).strip().split(",")[0].strip().title()
        if len(loc.split()) <= 4:
            return loc

    # 3. Scan full text for any known UK place name
    text_lower = text.lower()
    for place in UK_PLACES:
        if place.lower() in text_lower:
            return place

    return ""


def get_nearby_location(prospect_loc: str) -> str:
    """
    Given the prospect's location, return a different nearby UK place.
    """
    import random

    if not prospect_loc:
        return "the next county"

    # Direct lookup
    neighbours = UK_NEIGHBOURS.get(prospect_loc)
    if neighbours:
        return random.choice(neighbours)

    # Try town → county mapping
    county = TOWN_TO_COUNTY.get(prospect_loc)
    if county:
        neighbours = UK_NEIGHBOURS.get(county)
        if neighbours:
            return random.choice(neighbours)

    # Partial key match (e.g. "North Yorkshire" matches "Yorkshire")
    for key, vals in UK_NEIGHBOURS.items():
        if prospect_loc.lower() in key.lower() or key.lower() in prospect_loc.lower():
            return random.choice(vals)

    # Nothing found — return a placeholder that still reads naturally
    return "the next county"


def extract_location_for_body(opener: str, text: str) -> str:
    """Returns a NEARBY location (not the prospect's own) to use in the DM body."""
    prospect_loc = extract_prospect_location(opener, text)
    nearby = get_nearby_location(prospect_loc)
    print(f"Prospect location: {prospect_loc!r} -> body location: {nearby!r}")
    return nearby


def detect_trade(text: str) -> str:
    """
    Detect the trade type from the page text.
    Returns a plural noun for use in the DM body (e.g. "roofers", "landscapers").
    """
    text_lower = text.lower()

    trade_signals = [
        (["roofing", "roofer", "flat roof", "pitched roof", "epdm", "felt roof",
          "fascia", "soffit", "guttering", "lead flashing", "chimney", "velux"], "roofers"),
        (["drone", "aerial", "uav", "uas", "thermal imaging", "aerial survey",
          "aerial photography", "aerial footage", "roof inspection drone"], "drone operators"),
        (["renovation", "refurb", "kitchen reno", "bathroom reno", "full house",
          "house renovation", "property renovation", "interior renovation"], "renovation contractors"),
        (["builder", "building contractor", "extension", "loft conversion",
          "groundwork", "brickwork", "blockwork", "underpinning", "footings",
          "new build", "steel frame", "construction"], "builders"),
        (["landscap", "garden", "paving", "driveway", "patio", "turf", "decking",
          "fencing", "artificial grass", "tree work", "hedging", "planting",
          "block pav", "resin", "sandstone"], "landscapers"),
    ]

    scores = {}
    for keywords, trade in trade_signals:
        score = sum(1 for kw in keywords if kw in text_lower)
        if score > 0:
            scores[trade] = score

    if scores:
        best = max(scores, key=scores.get)
        print(f"Trade detected: {best} (scores: {scores})")
        return best

    return "tradespeople"


def detect_landscaper_type(text: str) -> str:
    """
    Sub-detect within landscapers: driveway, patio, or general garden.
    Used to pick the right 'my mate' body copy.
    """
    t = text.lower()
    driveway_kw = ["driveway", "block pav", "resin bound", "resin driveway",
                   "tarmac", "imprinted concrete", "pattern imprint", "tegula", "sett",
                   "monoblock", "cobble"]
    patio_kw    = ["patio", "porcelain", "indian sandstone", "limestone", "paving slab",
                   "decking", "composite deck", "sleeper", "sandstone", "flags", "flagstone"]

    driveway_score = sum(1 for kw in driveway_kw if kw in t)
    patio_score    = sum(1 for kw in patio_kw    if kw in t)

    if driveway_score > patio_score:
        sub = "driveway"
    elif patio_score > 0:
        sub = "patio"
    else:
        sub = "garden"
    print(f"Landscaper sub-type: {sub} (driveway={driveway_score}, patio={patio_score})")
    return sub


def extract_company_name(text: str) -> str:
    """
    Extract company name from scraped Facebook page text.
    The first meaningful line of page content is usually the page/company name.
    """
    if not text:
        return ""
    working = text
    if "--- PAGE TEXT ---" in working:
        working = working.split("--- PAGE TEXT ---", 1)[1]
    elif "--- TRADE TYPE ---" in working:
        after = working.split("--- TRADE TYPE ---", 1)[1].strip()
        lines = after.split("\n")
        working = "\n".join(lines[1:]) if len(lines) > 1 else ""
    for h in ("--- LISTED CITY ---", "--- USER NOTES ---", "--- DETECTED FIRST NAME ---"):
        if h in working:
            working = working.split(h, 1)[0]
    for line in working.strip().split("\n"):
        line = line.strip()
        if line and len(line) >= 3 and not line.startswith("---") and not line.startswith("http"):
            return line
    return ""


def _td_section(text_data: str, header: str) -> str:
    """Pull the first non-empty line after a --- HEADER --- marker."""
    marker = f"--- {header} ---"
    for line in text_data.split("\n"):
        if line.startswith(marker):
            idx = text_data.find(line) + len(line)
            rest = text_data[idx:].strip().split("\n")
            return rest[0].strip() if rest else ""
    return ""


def _pitch_for(niche: str):
    """(trade_noun, job_noun, example_questions) for the pitch body, tailored
    per niche so it reads bespoke. Matches on KEYWORDS contained in the value so
    it's robust to column phrasing ("HVAC / Air Con", "Gas Boiler", etc.).
    Unknown ("") stays NEUTRAL — never assumes a trade that could be wrong."""
    n = (niche or "").strip().lower()
    if not n:
        return ("trades", "a job", "size of the job, what's involved")
    has = lambda *words: any(w in n for w in words)
    # Order matters: bathroom first (a "Bathroom Renovation" value also contains
    # "renovation"); air con before heating; etc.
    if has("bathroom", "wet room", "ensuite"):
        return ("bathroom fitters", "a new bathroom", "full refit, wet room, retile")
    if has("hvac", "air con", "aircon", "air conditioning", "refrigeration", "cooling"):
        return ("air con engineers", "a new system", "one room, multi-split, or the whole place")
    if has("boiler", "gas", "heating", "central heat"):
        return ("heating engineers", "a new boiler", "combi swap, system upgrade, or just a repair")
    if has("plumb"):
        return ("plumbers", "a job", "a repair, a new install, or a full job")
    if has("renovation", "refurb", "general build", "builder", "extension", "remodel"):
        return ("renovation specialists", "a renovation", "how big a job, what's involved")
    # any other recognised trade (roofing, tiling, joinery, electrical, etc.)
    return (f"{n} businesses", "a job", "size of the job, what's involved")


def pick_body(trade: str, text_data: str, city_hint: str = "", company_name: str = "") -> str:
    """
    English demo-pitch body. Niche-tailored so each message reads bespoke.
    The `trade` arg is kept for signature compatibility but the niche is
    detected here via the name-first + text-scoring classifier.
    """
    # Always prefer the clean Excel name injected into text_data
    # over whatever extract_company_name() scraped from the page
    excel_name = _td_section(text_data, "COMPANY NAME")
    if excel_name:
        company_name = excel_name
    # Niche: prefer the Excel "type / niche" column (injected as TRADE TYPE);
    # fall back to the name/text classifier when that column is empty.
    excel_niche = _td_section(text_data, "TRADE TYPE")
    niche = excel_niche or _en_detect_niche(text_data, company_name)
    trade_noun, job, examples = _pitch_for(niche)

    who = _clean_company_name(company_name) or "your business"

    # Rotate the wording so two leads in the same niche never get a textually
    # identical body (kills Facebook's duplicate-content flag and the "template"
    # read if two owners ever compare notes). The meaning, the honest "student"
    # anchor and the "Can I send it over?" closer stay constant — only phrasing
    # moves. Same approach as pick_stripped_message().
    intro = random.choice([
        f"I'm a student studying AI, building quoting chatbots for {trade_noun}, "
        f"and I've already set one up for {who}, free. Just after honest feedback.",
        f"I'm a student getting into AI, I build quoting chatbots for {trade_noun}, "
        f"and I've put one together for {who} already, free. Just after honest feedback.",
        f"I'm a student learning AI, building quote bots for {trade_noun}, "
        f"and I've already made one for {who}, no charge. Just after some honest feedback.",
    ])
    how = random.choice([
        f"Someone messages asking what {job}'d cost, it asks a couple of quick questions "
        f"({examples}) and gives them a rough ballpark, then sends you their details.",
        f"When someone asks what {job}'d cost, it runs them through a couple of quick "
        f"questions ({examples}), gives a rough ballpark, then passes you their details.",
        f"A customer asks what {job}'d cost, it asks them a couple of quick things "
        f"({examples}), gives a rough price on the spot, then sends their details over to you.",
    ])
    benefit = random.choice([
        "So you skip the time wasters and only quote the serious ones.",
        "So you stop wasting time on tyre-kickers and only quote the serious ones.",
        "So the time wasters filter themselves out and you only quote the ones who mean it.",
    ])
    pitch = f"{intro}\n\n{how} {benefit}\n\nCan I send it over?\n\nCheers"

    return "\n" + pitch


# ── Owner email composer (for leads whose contact is an email) ─────────────────
EMAIL_SIGNOFF = "Milan"

def _email_first_name(owner_name: str) -> str:
    """First name for the email greeting, taken from the 'Owner Name' column.
    Returns '' when the value doesn't look like a person's name, so we fall back
    to a plain 'Hi,' rather than greeting a company by its trading name."""
    o = (owner_name or "").strip()
    if not o:
        return ""
    first = o.split()[0]
    bad = {"the", "a", "gas", "air", "ltd", "uk"}
    if first.isalpha() and len(first) >= 2 and first.lower() not in bad:
        return first[:1].upper() + first[1:]
    return ""

def _email_subject(company_name: str, niche: str = "", first: str = "") -> str:
    """Short, lowercase, personalised subject — the format the open-rate data backs
    (Gong's 85M-email study + Belkins 2025): all-lowercase beats title case ~15-20%,
    1-4 words wins, a first name lifts opens ~22% (the strongest lever), company name
    ~18%, and curiosity/question framing adds more. Zero selling words (no 'free').
    These lean on curiosity ('made you something', 'before they even ring') and a
    real personal hook ('saw your work') rather than the generic 'quick one'.
    Rotated so a batch isn't identical; falls back to the business name with no first
    name."""
    who = _clean_company_name(company_name) or "your business"
    # Short niche word for a relevant variant: "a new bathroom" -> "bathroom".
    _, job, _ex = _pitch_for(niche)
    thing = (job or "").replace("a new ", "").replace("a ", "").strip() or "job"
    if first:
        f = first.lower()   # all-lowercase reads like a colleague, not marketing
        return random.choice([
            f"{f}, made you something",
            f"saw your work, {f}",
            f"{f}, before they even ring",
            f"{f}, your {thing} enquiries",
        ])
    return random.choice([
        f"made something for {who}",
        f"saw your page, {who}",
        f"{who}, before they even ring",
    ])

def _clean_fb_title(title: str) -> str:
    """Pull a usable business name out of a Facebook page <title>. FB titles look
    like 'Steve Jones Plumbing | Facebook'. Crucially this must survive a Chrome
    that is logged into Facebook in ANOTHER language: the body text then contains
    UI chrome like 'Az olvasatlan értesítések száma' (Hungarian for 'unread
    notifications'), which the old body-text extractor wrongly used as the company
    name. We take the part before the separator and reject any FB/notification
    UI noise."""
    t = (title or "").strip()
    if not t:
        return ""
    for sep in ("|", "–", "—", " - ", "/", "·", "•"):
        if sep in t:
            t = t.split(sep)[0].strip()
            break
    low = t.lower()
    noise = ("facebook", "messenger", "értesít", "olvasatlan", "notification",
             "unread", "home", "kezdőlap", "log in", "bejelentkez")
    if not t or len(t) > 60 or any(n in low for n in noise):
        return ""
    return t

def _build_email(opener: str, page_text: str, lead: dict, page_title: str = "") -> tuple:
    """Compose the owner email as (subject, body).

    This is a COLD email to a busy tradesman who has plenty of competition, so it
    earns its place: one genuine line proving we looked at their page, then it
    leads with what's in it for THEM (instant prices filter time-wasters; serious
    leads arrive with budget + details; they reply before a competitor does). It
    does NOT explain how the tool works step by step — owners don't care. The
    "I already built one on your branding, want the link?" close makes it
    believable that a real thing exists, and the student framing explains the
    'free'. Segments are rotated so a batch isn't textually identical.

    Business name comes from the page <title> (clean), else 'your business' — we
    never scrape it from body text, which on a non-English FB session is noise."""
    owner = lead.get("name", "")
    city  = lead.get("city", "")
    niche = lead.get("niche", "")

    td = page_text or ""
    if niche:
        td = f"--- TRADE TYPE ---\n{niche}\n\n" + td
    if city:
        td += f"\n\n--- LISTED CITY ---\n{city}"

    company = _clean_fb_title(page_title) or "your business"
    niche_resolved = niche or _en_detect_niche(td, company)
    _, job, _examples = _pitch_for(niche_resolved)   # e.g. "a new bathroom"

    first = _email_first_name(owner)
    greet = f"Hey {first}," if first else "Hey,"

    # One genuine personal line (from their page), kept to a single sentence.
    op = (opener or "").strip()
    if op:
        op = op[:1].upper() + op[1:]
        if not op.endswith((".", "!", "?")):
            op += "."

    # LEAD with the honest student / case-study framing. Goal is real examples,
    # not money — and crucially never the word "free" (a classic spam-filter
    # trigger). This disarms the "what's the catch" before the pitch.
    frame = random.choice([
        ("I'm a student putting together instant-quoting tools for trades. Not after anything, "
         "I just want a couple of real businesses using one so I've got something to show."),
        ("I'm a student building instant-quoting tools for trades. Not after anything, I just "
         "want a couple of real businesses using one so I've got something to show."),
        ("I'm a student making instant-quoting tools for trades. Not chasing anything, I just "
         "want a couple of real businesses using one so I've got something to show."),
    ])

    # The benefit, led by filtering time-wasters (your chosen angle), niche-aware.
    benefit = random.choice([
        (f"It gives people a price for {job} before they even ring, so the time wasters drop "
         f"off and the serious ones land with their budget and details already filled in."),
        (f"It prices {job} for people before they even ring, so the time wasters drop off and "
         f"the serious ones come through with their budget and details already filled in."),
        (f"It gives people a price for {job} up front, so the time wasters drop off and only "
         f"the serious ones reach you, budget and details already filled in."),
    ])

    # Believable proof (a real thing exists, on their branding) + the ask, now
    # offering a channel choice (email here, or WhatsApp).
    close = random.choice([
        "I've built one on your branding already. Should I send it here or on WhatsApp?",
        "I've already built one on your branding. Want it here or on WhatsApp?",
        "Built one on your branding already. Shall I send it here or over WhatsApp?",
    ])

    blocks = [greet]
    if op:
        blocks.append(op)
    blocks += [frame, benefit, close, f"Cheers,\n{EMAIL_SIGNOFF}"]
    body = "\n\n".join(blocks)
    return _email_subject(company, niche, first), tidy_message(body)


def pick_stripped_message(text_data: str, name_hint: str = "", company_name: str = "") -> str:
    """Direct, honest A/B variant — no personalised opener, no vision call.
    'Hi [name], I'm a student building AI quote bots for [trade] and I've put a
    sample one together for [company]. No pressure, thought I'd just show you
    what it does rather than explain it. Can I send it over?'
    Leads with the honest "I'm a student" framing (disarming, not vendor-y).
    Closer never rotates. Currently the channel-neutral "Can I send it over?"
    (deliver the demo link on Messenger). Switch to "Can I send it here or would
    WhatsApp be better?" once WhatsApp is confirmed working.
    Niche comes from the Excel TRADE TYPE column first, else the name/text
    classifier — the SAME source the full body uses, so it stays niche-accurate."""
    excel_name = _td_section(text_data, "COMPANY NAME")
    if excel_name:
        company_name = excel_name
    excel_niche = _td_section(text_data, "TRADE TYPE")
    niche = excel_niche or _en_detect_niche(text_data, company_name)
    trade_noun, _job, _ex = _pitch_for(niche)

    # Straight in — no hollow pleasantries ("hope you're well"). They don't care
    # about that; brevity and getting to the point is the warmth here.
    greet = f"Hi {name_hint}, " if name_hint else "Hi, "
    who = _clean_company_name(company_name) or "your business"
    if name_hint and who.lower() == name_hint.lower():
        who = "you"   # sole trader: name == company, avoid "Hi Pierce ... for Pierce"

    # Rotate natural wordings of the middle so no two messages are textually
    # identical (kills Facebook's duplicate-content spam flag) while keeping the
    # exact meaning and tone. The "I'm a student" anchor and the "Can I send it
    # over?" closer never rotate out.
    build = random.choice([
        f"I'm a student building AI quote bots for {trade_noun}",
        f"I'm a student getting into AI, building quote bots for {trade_noun}",
        f"I'm a student learning AI, I build quote bots for {trade_noun}",
    ])
    sample = random.choice([
        f"and I've put a sample one together for {who}",
        f"and I've built a sample for {who}",
        f"and I've made a sample one for {who}",
    ])
    show = random.choice([
        "No pressure, thought I'd just show you what it does rather than explain it.",
        "No pressure, figured it's easier to show you than explain it.",
        "No pressure, thought it'd be easier to just show you what it does.",
    ])
    return f"{greet}{build} {sample}. {show} Can I send it over?"


EXTRACT_PROMPT = """\
Look at this trade business page text carefully. List ONLY details you can actually read. Do not guess.

Output a short bullet list with ONLY these if present:
- Trade type (landscaper / roofer / builder / renovation contractor / drone operator)
- MOST RECENT POST: job type, material, colour/finish, any location in the caption (full place name, no abbreviations or postcodes)
- Other post locations: any town/city names in post captions (full names only)
- Profile location: the About/bio location (for body message use)
- Owner first name (ONLY if in the company name, email, or About — not guessed)
- Anything unusual: massive job, niche material, striking before/after, unusual service

Do not include anything not clearly written in the text. Bullet list only, nothing else."""


_FALLBACK_TEMPLATES = {
    "roofers": [
        "Came across your page over in {loc}, fair bit of roofing work on there.",
        "Clocked your page over in {loc}, looks like you keep busy with the roofing.",
    ],
    "landscapers": [
        "Came across your page over in {loc}, fair bit of landscaping work on there.",
        "Clocked your page over in {loc}, looks like you keep busy with the landscaping.",
    ],
    "builders": [
        "Came across your page over in {loc}, fair bit of building work on there.",
        "Clocked your page over in {loc}, some solid building work on there.",
    ],
    "renovation contractors": [
        "Came across your page over in {loc}, decent renovation work on there.",
    ],
    "drone operators": [
        "Came across your page over in {loc}, decent aerial work on there.",
    ],
}
_FALLBACK_LOC_ONLY = [
    "Came across your page over in {loc}, solid work on there.",
    "Clocked your page over in {loc}, looks like you keep busy.",
]
_FALLBACK_NOTHING = "Came across your page, solid work on there."


def copy_to_clipboard(text: str) -> bool:
    """Copy text to Windows clipboard. Returns True if successful."""
    if not _CLIPBOARD_OK:
        return False
    try:
        _pyperclip.copy(text)
        print("  DM copied to clipboard.")
        return True
    except Exception as e:
        print(f"  Clipboard copy failed: {e}")
        return False


def patch_fallback_opener(opener: str, text_data: str, city_hint: str = "") -> str:
    """
    If the model used the generic fallback, normalise it to the plain
    'Came across your page in [location].' form (mirror of the HU fallback).
    Runs AFTER call_gemini() — Python safety net.
    """
    # Only touch it if it's the generic fallback
    if "came across your page" not in opener.lower() and "clocked your page" not in opener.lower():
        return opener

    loc = (_td_section(text_data, "LISTED CITY")
           or extract_prospect_location("", text_data)
           or city_hint)
    result = f"Came across your page in {loc}." if loc else "Came across your page."
    print(f"Fallback normalised -> {result!r}")
    return result


# ── Niche resolution + prompt builder + opener post-validation ───────────────
def _canon_niche(value: str) -> str:
    """Map a free-form niche/column value to the canonical niche key used by the
    opener instructions and the contradiction guard. Keyword-based so it copes
    with 'Bathroom Renovation', 'Gas Boiler', 'HVAC / Air Con', etc."""
    n = (value or "").strip().lower()
    if not n:
        return ""
    has = lambda *w: any(x in n for x in w)
    if has("bathroom", "wet room", "ensuite"):            return "bathroom renovation"
    if has("hvac", "air con", "aircon", "air conditioning", "refrigeration", "cooling"):
        return "air conditioning"
    if has("boiler", "gas", "heating", "central heat"):   return "heating"
    if has("plumb"):                                      return "plumbing"
    if has("window", "glazing", "conservatory", "upvc"):  return "windows and doors"
    if has("roof"):                                       return "roofing"
    if has("solar", "photovolt"):                         return "solar"
    if has("paving", "driveway", "block pav", "resin"):   return "paving"
    if has("tiling", "tiler"):                            return "tiling"
    if has("render", "ewi", "k rend"):                    return "external wall insulation"
    if has("landscap", "garden"):                         return "landscaping"
    if has("decking"):                                    return "decking"
    if has("joinery", "joiner", "carpenter", "kitchen"):  return "joinery"
    if has("electric"):                                   return "electrical"
    if has("renovation", "refurb", "general build", "builder", "extension"):
        return "general building"
    return n


def _resolve_niche(text_data: str) -> str:
    """
    Resolve the niche the SAME way the body does: the Excel TRADE TYPE column is
    authoritative (so the opener and the pitch never disagree), then name-first,
    then full text-scoring. Result is a canonical key.
    """
    excel_niche = _td_section(text_data, "TRADE TYPE")
    if excel_niche:
        return _canon_niche(excel_niche)
    company_name = _td_section(text_data, "COMPANY NAME")
    if company_name:
        niche = _en_detect_niche_from_name(company_name)
        if niche:
            return niche
    return _en_detect_niche(text_data, company_name)


def _en_prompt_builder(full_prompt: str, text_data: str) -> str:
    """Inject per-niche context before the vision call."""
    niche = _resolve_niche(text_data)
    # Unknown ("") still gets the general-builder guidance ("react to whatever
    # trade work you can see") so the vision opener stays useful.
    instr = (_EN_NICHE_OPENER_INSTRUCTIONS.get(niche)
             or _EN_NICHE_OPENER_INSTRUCTIONS.get("general building"))
    if instr:
        print(f"  Prompt builder: niche='{niche or '(unknown)'}' — injecting instructions")
        return full_prompt + f"\n\n{instr}"
    return full_prompt


# Generic trade / legal words — a pair made of ONLY these isn't "the company
# name" (e.g. "bathroom renovation"), so it shouldn't trigger the name guard.
_NAME_GUARD_GENERIC = {
    "the", "and", "co", "ltd", "limited", "llp", "uk", "group", "services",
    "service", "bathroom", "bathrooms", "renovation", "renovations",
    "refurbishment", "refurbishments", "fitter", "fitters", "fitting",
    "installations", "installation", "kitchen", "kitchens", "plumbing",
    "heating", "gas", "boiler", "tiling", "design", "designs", "builders",
    "building", "construction", "company", "air", "con", "conditioning",
    "hvac", "interiors", "interior", "property", "maintenance", "solutions",
    "contractors", "specialist", "specialists",
}


def _opener_repeats_company_name(opener: str, company: str) -> bool:
    """True if the opener embeds 2+ consecutive words of the company name where
    at least one is distinctive (e.g. 'That Sussex Bathrooms installation...').
    A single shared word like 'bathroom' is fine; a generic-only pair like
    'bathroom renovation' doesn't count."""
    if not opener or not company:
        return False
    clean = lambda s: re.sub(r"[^a-z0-9 ]", " ", s.lower()).split()
    o, c = clean(opener), clean(company)
    for i in range(len(c) - 1):
        pair = c[i:i + 2]
        if all(t in _NAME_GUARD_GENERIC for t in pair):
            continue  # generic-only pair, not a real name match
        for j in range(len(o) - 1):
            if o[j:j + 2] == pair:
                return True
    return False


def _en_opener_postprocess(opener: str, text_data: str) -> str:
    """Reject opener if it contradicts the detected niche OR embeds the company
    name; fall back to the city line."""
    niche = _resolve_niche(text_data)
    company = _td_section(text_data, "COMPANY NAME")
    reason = ""
    if not _en_opener_matches_niche(opener, niche):
        reason = f"contradicts niche '{niche}'"
    elif _opener_repeats_company_name(opener, company):
        reason = "embeds the company name"
    if reason:
        city = _td_section(text_data, "LISTED CITY")
        fallback = (f"Came across your page in {city}."
                    if city else "Came across your page.")
        print(f"  Opener REJECTED ({reason}): '{opener[:60]}' -> '{fallback}'")
        return fallback
    return opener


CUSTOM_PROMPT_BUILDER     = _en_prompt_builder
CUSTOM_OPENER_POSTPROCESS = _en_opener_postprocess


# ── Groq call ────────────────────────────────────────────────────────────────
def _run_groq(client, model: str, content, max_tokens: int = 500, temperature: float = 0.7) -> str:
    import time
    resp = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": content}],
        max_tokens=max_tokens,
        temperature=temperature,
    )
    return resp.choices[0].message.content.strip()


def _try_models(client, attempts, max_tokens=500, temperature=0.7) -> str:
    import time
    last_error = None
    for model, content in attempts:
        for attempt in range(3):
            try:
                result = _run_groq(client, model, content, max_tokens, temperature)
                return result
            except Exception as e:
                msg = str(e)
                last_error = msg
                skip_keywords = ("404", "not found", "decommissioned", "model_not_found", "no longer supported")
                if any(k in msg.lower() for k in skip_keywords):
                    print(f"{model} not available, skipping...")
                    break
                if attempt < 1:
                    time.sleep(2)
                    continue
                print(f"{model} failed ({msg[:80]}), trying next...")
                break
    raise RuntimeError(f"All Groq models failed. Last error: {last_error}")


def call_gemini(text_data: str = "", images: list = None, prompt: str = None) -> str:
    client = Groq(api_key=API_KEY)
    if prompt is None:
        prompt = PROMPT

    # ── Step 1: extract clean facts from page text (fast, text-only) ──────────
    facts = text_data
    if text_data and len(text_data.strip()) > 200:
        try:
            extract_prompt = EXTRACT_PROMPT + f"\n\nPAGE TEXT:\n{text_data[:4000]}"
            facts = _try_models(client,
                [(m, extract_prompt) for m in FACTS_MODELS],
                max_tokens=250, temperature=0.2)
            print(f"Extracted facts:\n{facts}\n---")
        except Exception:
            facts = text_data  # fall back to raw text

    # ── Step 2: write opener from clean facts ─────────────────────────────────
    context_str = f"\n\nPAGE FACTS:\n{facts}" if facts else ""
    full_prompt = prompt + context_str
    if CUSTOM_PROMPT_BUILDER:
        full_prompt = CUSTOM_PROMPT_BUILDER(full_prompt, text_data) or full_prompt

    write_attempts = []
    if images:
        vision_content = []
        for img_bytes, mime in images:
            b64 = base64.b64encode(img_bytes).decode()
            vision_content.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}})
        vision_content.append({"type": "text", "text": full_prompt})
        write_attempts += [(m, vision_content) for m in VISION_MODELS]
    write_attempts += [(m, full_prompt) for m in TEXT_MODELS]

    last_error = None
    for model, content in write_attempts:
        for attempt in range(3):
            try:
                import time
                result = _run_groq(client, model, content, max_tokens=400, temperature=0.92)
                if re.search(r"\b\d[\d,]*\s*(followers?|likes?|reviews?|ratings?|fans?)", result, re.I):
                    print(f"Output contained forbidden metric, retrying...")
                    continue
                if CUSTOM_OPENER_POSTPROCESS:
                    result = CUSTOM_OPENER_POSTPROCESS(result, text_data) or result
                return result
            except Exception as e:
                msg = str(e)
                last_error = msg
                skip_keywords = ("404", "not found", "decommissioned", "model_not_found", "no longer supported")
                if any(k in msg.lower() for k in skip_keywords):
                    print(f"{model} not available, skipping...")
                    break
                if attempt < 1:
                    time.sleep(2)
                    continue
                print(f"{model} failed ({msg[:80]}), trying next...")
                break
    raise RuntimeError(f"All Groq models failed. Last error: {last_error}")

def fix_grammar(text: str) -> str:
    """Run the final DM through Groq to fix any grammatical errors. Hungarian-aware."""
    if not GRAMMAR_CHECK or not text.strip():
        return text
    try:
        client = Groq(api_key=API_KEY)
        prompt = (
            "Te egy magyar nyelvhelyességi ellenőrző vagy. "
            "Javítsd ki a következő szöveg összes nyelvtani hibáját. "
            "Fontos szabályok: "
            "1. Ha egy kérdés 'Ti,' szóval kezdődik vesszővel, töröld a 'Ti,' részt. "
            "2. Ne változtasd meg a jelentést, a stílust vagy a hangnemet. "
            "3. Ne adj hozzá semmit, ne magyarázz. Csak a javított szöveget add vissza.\n\n"
            f"{text}"
        )
        result = _run_groq(client, "llama-3.3-70b-versatile", prompt, max_tokens=300, temperature=0.1)
        print(f"Grammar fix: {text!r} -> {result!r}")
        return result
    except Exception as e:
        print(f"Grammar check failed, using original: {e}")
        return text


def tidy_message(text: str) -> str:
    """Instant, local cleanup of the assembled DM. Replaces the old LLM
    grammar pass (which was a HUNGARIAN checker being run on ENGLISH copy —
    wasted a round-trip per lead and could mangle good lines). Cosmetic only:
    collapse repeated spaces, drop space-before-punctuation, trim line ends.
    Never changes wording. Preserves the paragraph breaks (blank lines)."""
    if not text:
        return text
    out = []
    for line in text.split("\n"):
        line = re.sub(r"[ \t]{2,}", " ", line)        # collapse runs of spaces/tabs
        line = re.sub(r"[ \t]+([,.!?;:])", r"\1", line)  # no space before punctuation
        out.append(line.rstrip())
    return "\n".join(out).strip()


# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("dm.html", default_excel=DEFAULT_EXCEL, variant=MESSAGE_VARIANT)

@app.route("/toggle_variant", methods=["POST"])
def toggle_variant():
    """Flip the A/B message variant (full <-> stripped) and persist it."""
    global MESSAGE_VARIANT
    MESSAGE_VARIANT = "stripped" if MESSAGE_VARIANT == "full" else "full"
    _save_variant(MESSAGE_VARIANT)
    print(f"Message variant -> {MESSAGE_VARIANT}")
    return jsonify({"variant": MESSAGE_VARIANT})

@app.route("/generate", methods=["POST"])
def generate():
    try:
        if not API_KEY:
            return jsonify({"error": "GROQ_API_KEY not set in .env"}), 500

        images    = []
        text_data = ""
        warning   = ""
        name_hint = ""

        # Manual image upload (screenshot pasted in the app)
        if "image" in request.files and request.files["image"].filename:
            f = request.files["image"]
            images.append((f.read(), f.mimetype or "image/jpeg"))
            print(f"Manual image uploaded: {f.mimetype}, {len(images[0][0])} bytes")

        # Manual notes
        notes = request.form.get("notes", "").strip()
        if notes:
            text_data += f"\n--- USER NOTES ---\n{notes}"

        # Capture from Chrome (primary method — only if no manual image)
        if not images:
            hint_url = request.form.get("url", "").strip()
            captured_images, captured_text = capture_profile_tab(hint_url)
            if captured_images or captured_text:
                images = captured_images  # may be empty if text was rich — that's fine
                if captured_text:
                    name_hint = extract_first_name(captured_text)
                    if name_hint:
                        print(f"Name detected: {name_hint}")
                        text_data += f"\n--- DETECTED FIRST NAME ---\n{name_hint}"
                    text_data += f"\n--- PAGE TEXT ---\n{captured_text}"
                print(f"Captured {len(images)} screenshots, {len(captured_text)} chars text from Chrome")
            else:
                warning = (
                    "No Facebook/Instagram profile tab found in Chrome. "
                    "Open the profile page in Chrome first, then click Generate."
                )
                if not text_data.strip():
                    text_data = "--- NO PAGE DATA --- Do NOT invent any details."

        if not images and not text_data.strip() and not notes:
            return jsonify({
                "error": "Open the profile in Chrome first, then click Generate. "
                         "Or paste a screenshot / type profile info."
            }), 400

        if MESSAGE_VARIANT == "stripped":
            # Direct one-line variant — no opener, so skip the vision call.
            dm = tidy_message(pick_stripped_message(
                text_data, name_hint, extract_company_name(text_data)))
            dms = [dm]
        else:
            company = extract_company_name(text_data)
            trade   = detect_trade(text_data)

            def _assemble(opener: str) -> str:
                opener = opener.rstrip(".!,") + "."
                opener = patch_fallback_opener(opener, text_data)
                if name_hint:
                    g = GREETING_NAME_FMT.format(prefix=GREETING_PREFIX, name=name_hint)
                    opener = g + (opener if g.rstrip().endswith(".")
                                  else opener[0].lower() + opener[1:])
                elif ALWAYS_GREET:
                    opener = GREETING_SOLO_FMT.format(prefix=GREETING_PREFIX) + opener
                body = pick_body(trade, text_data, company_name=company)
                return tidy_message((opener + " " + body).strip())

            # Generate up to 3 candidates so you pick the best read instead of
            # regenerating by hand. Vision temp 0.92 + rotated body keep them
            # distinct; we dedup on the opener so identical fallbacks collapse.
            dms, seen = [], set()
            # First candidate — let a hard failure bubble up with its real error.
            first = call_gemini(text_data=text_data, images=images if images else None)
            seen.add(re.sub(r"\s+", " ", first.strip().lower()))
            dms.append(_assemble(first))
            # Two more best-effort candidates; ignore transient failures.
            for _ in range(2):
                try:
                    raw = call_gemini(text_data=text_data, images=images if images else None)
                except Exception:
                    break
                key = re.sub(r"\s+", " ", raw.strip().lower())
                if key in seen:
                    continue
                seen.add(key)
                dms.append(_assemble(raw))
            dm = dms[0]
        clipboard = copy_to_clipboard(dm)
        return jsonify({"dm": dm, "dms": dms, "warning": warning, "clipboard": clipboard})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/followup", methods=["POST"])
def followup():
    data  = request.get_json(force=True)
    trade = (data.get("trade") or "quotes").lower().strip()
    touch = int(data.get("touch", 1)) - 1  # 0-indexed

    sequence = FOLLOWUPS.get(trade) or FOLLOWUPS["quotes"]
    touch = max(0, min(touch, len(sequence) - 1))
    msg   = sequence[touch]

    copy_to_clipboard(msg)
    return jsonify({"message": msg})


@app.route("/generate_voicenote", methods=["POST"])
def generate_voicenote():
    try:
        if not API_KEY:
            return jsonify({"error": "GROQ_API_KEY not set in .env"}), 500

        images    = []
        text_data = ""
        warning   = ""

        if "image" in request.files and request.files["image"].filename:
            f = request.files["image"]
            images.append((f.read(), f.mimetype or "image/jpeg"))

        notes = request.form.get("notes", "").strip()
        if notes:
            text_data += f"\n--- USER NOTES ---\n{notes}"

        if not images:
            hint_url = request.form.get("url", "").strip()
            captured_images, captured_text = capture_profile_tab(hint_url)
            if captured_images:
                images = captured_images
                if captured_text:
                    name_hint = extract_first_name(captured_text)
                    if name_hint:
                        text_data += f"\n--- DETECTED FIRST NAME ---\n{name_hint}"
                    text_data += f"\n--- PAGE TEXT ---\n{captured_text}"
            else:
                warning = (
                    "No Facebook/Instagram profile tab found in Chrome. "
                    "Open the profile page in Chrome first, then click Generate."
                )
                if not text_data.strip():
                    text_data = "--- NO PAGE DATA --- Do NOT invent any details."

        if not images and not notes:
            return jsonify({
                "error": "Open the profile in Chrome first, then click Generate. "
                         "Or paste a screenshot / type profile info."
            }), 400

        script = call_gemini(text_data=text_data, images=images if images else None, prompt=VOICE_NOTE_PROMPT)
        return jsonify({"script": script, "warning": warning})

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ── Auto Sender state ──────────────────────────────────────────────────────────
import openpyxl as _xl
from pathlib import Path as _Path

_auto_state = {
    "excel_path":    None,
    "current_row":   None,
    "sent_today":    0,
    "wb":            None,
    "ws":            None,
    "cols":          {},
}

# Separate state for the Owner Emails section so it never interferes with the
# Auto Sender (DM) flow — you can load a different list in each.
_email_state = {
    "excel_path":    None,
    "current_row":   None,
    "sent_today":    0,
    "wb":            None,
    "ws":            None,
    "cols":          {},
}

DAILY_CAP = 30

def _load_excel(path: str):
    wb = _xl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in ws[1]]

    created = []
    def _find_or_create(match, label):
        """Return the index of the first header matching `match(header)`,
        creating a new column titled `label` at the end if none exists."""
        idx = next((i for i, h in enumerate(headers) if match(h)), None)
        if idx is None:
            ws.cell(row=1, column=len(headers) + 1, value=label)
            headers.append(label.lower())
            idx = len(headers) - 1
            created.append(label)
        return idx

    # Follow-up tracking columns are created on first load if absent.
    # Match the SENT-tracking column only — NOT the lead-contact column. On the
    # newer list "Contact (Email or Personal FB)" also contains "contact", so the
    # old loose "contact" match hijacked the email column and overwrote it with
    # "Skip"/"Yes". Match "contacted"/"outreach status" instead.
    contacted_idx = _find_or_create(
        lambda h: "contacted" in h or "outreach" in h or h.strip() == "status",
        "Contacted?")
    sent_date_idx = _find_or_create(lambda h: "date" in h, "Sent Date")
    replied_idx   = _find_or_create(lambda h: "repl" in h, "Replied?")
    followups_idx = _find_or_create(lambda h: "follow" in h, "Follow-Ups")
    if created:
        wb.save(path)  # only write when we actually added a column

    cols = {
        "url":       next((i for i, h in enumerate(headers) if "url" in h or "facebook" in h), 1),
        "name":      next((i for i, h in enumerate(headers) if "company" in h or "name" in h), 0),
        "city":      next((i for i, h in enumerate(headers) if "city" in h or "location" in h), None),
        "niche":     next((i for i, h in enumerate(headers) if "niche" in h or "trade" in h), None),
        # The lead's primary contact (email OR personal FB) and its type. Used to
        # decide per lead whether we send a DM or write an owner email.
        "contact":      next((i for i, h in enumerate(headers) if "contact" in h and ("email" in h or "fb" in h or "personal" in h)), None),
        "contact_type": next((i for i, h in enumerate(headers) if "contact" in h and "type" in h), None),
        "company_fb":   next((i for i, h in enumerate(headers) if ("company" in h or "business" in h) and ("facebook" in h or "fb" in h or "page" in h)), None),
        "contacted": contacted_idx,
        "sent_date": sent_date_idx,
        "replied":   replied_idx,
        "followups": followups_idx,
    }
    return wb, ws, cols

def _save_wb():
    """Save the loaded workbook, retrying briefly if the file is locked
    (open in Excel, or mid OneDrive-sync). Raises a clear, actionable
    error if it stays locked so the UI can tell the user what to do."""
    import time as _t
    wb   = _auto_state.get("wb")
    path = _auto_state.get("excel_path")
    if wb is None or not path:
        return
    for attempt in range(6):
        try:
            wb.save(path)
            return
        except PermissionError:
            _t.sleep(0.4)
    raise PermissionError(
        "Can't write to the Excel file — it's open in Excel (or still syncing in "
        "OneDrive). Close the spreadsheet in Excel, wait for OneDrive to finish, "
        "then click again."
    )

def _today_str() -> str:
    from datetime import date
    return date.today().isoformat()

def _parse_date(v):
    """Coerce an Excel cell (datetime, date, or 'YYYY-MM-DD' string) to a date."""
    from datetime import datetime, date
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s[:19]).date()
    except Exception:
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").date()
        except Exception:
            return None

def _to_int(v):
    try:
        return int(float(str(v).strip()))
    except Exception:
        return 0

def _next_lead(ws, cols, skip_row=None):
    """Next uncontacted DM lead (Auto Sender). DM-only, unchanged: walks rows
    whose URL column is a real link. Email leads are handled by the separate
    Owner Emails section via _next_email_lead()."""
    for row in ws.iter_rows(min_row=2):
        excel_row = row[0].row
        if skip_row is not None and excel_row == skip_row:
            continue
        url = str(row[cols["url"]].value or "").strip()
        contacted = str(row[cols["contacted"]].value or "").strip().lower() if cols["contacted"] is not None else ""
        if not url.startswith("http"):
            continue  # skip blank or invalid URL rows silently
        if contacted not in ("yes", "y", "sent", "1", "true", "skip"):
            return {
                "row":   excel_row,
                "url":   url,
                "name":  str(row[cols["name"]].value or "").strip(),
                "city":  str(row[cols["city"]].value or "").strip()  if cols["city"]  is not None else "",
                "niche": str(row[cols["niche"]].value or "").strip() if cols["niche"] is not None else "",
            }
    return None


def _next_email_lead(ws, cols, skip_row=None):
    """Next uncontacted EMAIL lead (Owner Emails section). Only rows whose
    contact is an email address (or Contact Type says Email). Returns the email
    address plus the Company Facebook page to scan for a personal line."""
    for row in ws.iter_rows(min_row=2):
        excel_row = row[0].row
        if skip_row is not None and excel_row == skip_row:
            continue

        def _v(key):
            i = cols.get(key)
            return str(row[i].value or "").strip() if i is not None else ""

        if _v("contacted").lower() in ("yes", "y", "sent", "1", "true", "skip"):
            continue

        contact = _v("contact") or _v("url")
        ctype   = _v("contact_type").lower()
        if not (("@" in contact) or ("email" in ctype)):
            continue  # not an email lead — leave it for the DM flow

        company_fb = _v("company_fb")
        scan_url   = company_fb if company_fb.startswith("http") else ""
        return {
            "row":   excel_row,
            "email": contact,
            "url":   scan_url,   # Company FB page to scan ("" if none)
            "name":  _v("name"),
            "city":  _v("city"),
            "niche": _v("niche"),
        }
    return None

def _find_message_box(ctx):
    """Find Messenger input across all open tabs."""
    import time as _time
    _time.sleep(1.2)
    selectors = [
        '[aria-label="Message"]', '[aria-label="Aa"]',
        'div[contenteditable="true"][role="textbox"]',
        'div[contenteditable="true"]',
    ]
    for pg in reversed(list(ctx.pages)):
        if "facebook.com" in pg.url or "messenger.com" in pg.url:
            for sel in selectors:
                try:
                    el = pg.locator(sel).first
                    if el.is_visible(timeout=1200):
                        return pg, el
                except:
                    continue
    return None, None

def _click_message_btn(page):
    selectors = [
        # English Facebook
        '[aria-label="Send message"]', '[aria-label="Message"]',
        # Hungarian Facebook
        '[aria-label="Üzenet küldése"]', '[aria-label="Üzenet"]',
        # Generic fallbacks
        'a[href*="messenger.com/t/"]',
        'div[role="button"]:has-text("Message")',
        'div[role="button"]:has-text("Üzenet")',
        'a:has-text("Send message")', 'a:has-text("Message")',
        'a:has-text("Üzenet küldése")',
    ]
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=800):
                el.click()
                page.wait_for_timeout(2000)
                return True
        except:
            continue
    return False

@app.route("/auto_load", methods=["POST"])
def auto_load():
    """Load Excel and return first uncontacted lead info."""
    try:
        data = request.get_json()
        path = data.get("path", "").strip()

        # Auto-detect if no path given
        if not path:
            for f in _Path(__file__).parent.glob("*.xlsx"):
                nm = f.name.lower()
                if any(k in nm for k in ("lead", "dm", "contact", "prospect", "outreach")):
                    path = str(f); break
            if not path:
                files = list(_Path(__file__).parent.glob("*.xlsx"))
                path = str(files[0]) if files else ""

        if not path or not _Path(path).exists():
            return jsonify({"error": f"Excel file not found: {path}"}), 400

        wb, ws, cols = _load_excel(path)
        _auto_state.update({"excel_path": path, "wb": wb, "ws": ws, "cols": cols})

        lead = _next_lead(ws, cols)
        if not lead:
            return jsonify({"error": "No more leads — all done!"}), 200

        total = sum(1 for row in ws.iter_rows(min_row=2) if str(row[cols["url"]].value or "").strip())
        sent  = sum(1 for row in ws.iter_rows(min_row=2)
                    if cols["contacted"] is not None and
                    str(row[cols["contacted"]].value or "").strip().lower() in ("yes","y","sent","1","true"))

        return jsonify({"lead": lead, "total": total, "sent": sent})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/auto_next", methods=["POST"])
def auto_next():
    """Open next lead in Chrome, generate DM, type it into Facebook."""
    try:

        ws   = _auto_state["ws"]
        cols = _auto_state["cols"]
        if ws is None:
            return jsonify({"error": "Load your Excel file first."}), 400

        # If we still have a current_row that was never marked sent/skip,
        # mark it as Skip now so it never resurfaces as a duplicate.
        prev_row = _auto_state["current_row"]
        if prev_row is not None and cols["contacted"] is not None:
            ws.cell(row=prev_row, column=cols["contacted"] + 1, value="Skip")
            _save_wb()

        lead = _next_lead(ws, cols, skip_row=prev_row)
        if not lead:
            return jsonify({"error": "No more leads — all done!"}), 200

        _auto_state["current_row"] = lead["row"]

        # Open profile in Chrome
        from playwright.sync_api import sync_playwright
        import socket
        connected = False
        for _ in range(5):
            try:
                s = socket.create_connection(("127.0.0.1", CDP_PORT), timeout=2)
                s.close()
                connected = True
                break
            except:
                import time as _t2; _t2.sleep(1)
        if not connected:
            return jsonify({"error": "Chrome not found. Close all Chrome windows, double-click START.bat, wait for Chrome to fully open, then try again."}), 500

        with sync_playwright() as p:
            browser = p.chromium.connect_over_cdp(f"http://localhost:{CDP_PORT}")
            ctx     = browser.contexts[0]

            # Close any old Facebook profile tabs (not Messenger, not the app)
            for _pg in list(ctx.pages):
                _u = _pg.url.lower()
                if ("facebook.com/" in _u and
                        "messenger.com" not in _u and
                        "/messages" not in _u and
                        "localhost" not in _u):
                    try:
                        _pg.close()
                        print(f"  Closed old tab: {_u[:70]}")
                    except Exception:
                        pass

            page = ctx.new_page()
            try:
                page.goto(lead["url"], wait_until="domcontentloaded", timeout=45000)
            except Exception:
                try:
                    page.goto(lead["url"], wait_until="commit", timeout=30000)
                except Exception as e:
                    return jsonify({"error": f"Could not load page: {e}"}), 500
            page.wait_for_timeout(1000)   # settle after load (image-render wait below protects the screenshot)

            # Strip whatever overlay is dimming the page (a dialog or a full-screen
            # scrim). It dims the screenshot the vision model reads and can sit
            # invisibly on top of the Message button. Also logs every full-viewport
            # fixed/absolute layer to server.log so we can target it if it persists.
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass
            try:
                page.evaluate(r"""() => {
                    // 1) Persistent CSS so the login/cookie dialog stays hidden even
                    //    when React re-adds it (the one-time removal loses that race).
                    if (!document.getElementById('dm-overlay-killer')) {
                        const st = document.createElement('style');
                        st.id = 'dm-overlay-killer';
                        st.textContent =
                            '[role="dialog"]{display:none!important;}' +
                            'html,body{overflow:auto!important;}';
                        (document.head || document.documentElement).appendChild(st);
                    }
                    // 2) Strip full-viewport dark/semi-transparent backdrop scrims, and
                    //    keep doing so for a few seconds as Facebook re-renders.
                    const nuke = () => {
                        document.querySelectorAll('[role="dialog"]').forEach(el => el.remove());
                        document.querySelectorAll('body *').forEach(el => {
                            const s = getComputedStyle(el);
                            if (s.position !== 'fixed' && s.position !== 'absolute') return;
                            const r = el.getBoundingClientRect();
                            if (r.width < innerWidth * 0.8 || r.height < innerHeight * 0.6) return;
                            const bg = s.backgroundColor || '';
                            const op = parseFloat(s.opacity || '1');
                            const m  = bg.match(/rgba?\(([^)]+)\)/);
                            let a = 1; if (m) { const p = m[1].split(','); a = p.length > 3 ? parseFloat(p[3]) : 1; }
                            if ((a > 0 && a < 1) || (op > 0 && op < 1) || bg === 'rgb(0, 0, 0)') el.remove();
                        });
                        document.documentElement.style.overflow = 'auto';
                        document.body.style.overflow = 'auto';
                    };
                    nuke();
                    const obs = new MutationObserver(nuke);
                    obs.observe(document.documentElement, { childList: true, subtree: true });
                    setTimeout(() => obs.disconnect(), 6000);
                }""")
            except Exception:
                pass

            # Extract text
            try:
                text = page.evaluate("""() => {
                    ['script','style','noscript','svg','iframe']
                        .forEach(t => document.querySelectorAll(t).forEach(el => el.remove()));
                    return document.body ? document.body.innerText : '';
                }""")
                import re as _re
                text = _re.sub(r"\n{3,}", "\n\n", text).strip()
                text = _re.sub(r"(?im)^\s*[\d,\.]+\s*(followers?|likes?|reviews?|ratings?|check.?ins?|people follow|people like).*$", "", text)
                text = _re.sub(r"\n{3,}", "\n\n", text).strip()[:12000]
            except:
                text = ""

            text_data = text
            if lead["niche"]: text_data = f"--- TRADE TYPE ---\n{lead['niche']}\n\n" + text_data
            if lead["name"]:  text_data = f"--- COMPANY NAME ---\n{lead['name']}\n\n" + text_data
            if lead["city"]:  text_data += f"\n\n--- LISTED CITY ---\n{lead['city']}"

            # Extract name — only from the company name or a bio self-intro,
            # validated against FIRST_NAMES. No name found → plain "Hi." (no guess).
            name_hint = extract_first_name(text or "", lead.get("name", ""))

            if MESSAGE_VARIANT == "stripped":
                # Direct one-line variant — no opener, so skip the screenshot AND
                # the vision call entirely (faster, and the photo isn't used).
                dm = tidy_message(pick_stripped_message(
                    text_data, name_hint,
                    lead.get("name", "") or extract_company_name(text_data)))
            else:
                # Screenshot the first-post area so the vision model can see the actual job photos
                images = []
                try:
                    import time as _tw
                    page.evaluate("window.scrollTo(0, 850)")
                    _tw.sleep(1.4)   # wait for images to render
                    shot = page.screenshot(type="jpeg", quality=72)
                    images = [(shot, "image/jpeg")]
                    print(f"  First-post screenshot: {len(shot)} bytes")
                    page.evaluate("window.scrollTo(0, 0)")
                except Exception as _se:
                    print(f"  Screenshot skipped: {_se}")

                # Generate DM — pass screenshot so vision model sees the actual job
                opener = call_gemini(text_data=text_data, images=images if images else None)
                opener = opener.rstrip(".!,") + "."
                opener = patch_fallback_opener(opener, text_data, lead.get("city", ""))
                if name_hint:
                    g = GREETING_NAME_FMT.format(prefix=GREETING_PREFIX, name=name_hint)
                    opener = g + (opener if g.rstrip().endswith(".") else opener[0].lower() + opener[1:])
                elif ALWAYS_GREET:
                    opener = GREETING_SOLO_FMT.format(prefix=GREETING_PREFIX) + opener
                trade  = detect_trade(text_data)
                body   = pick_body(trade, text_data, lead.get("city", ""), lead.get("name", "") or extract_company_name(text_data))
                dm     = tidy_message((opener + " " + body).strip())

            # Copy DM to clipboard — always reliable
            clipboard = copy_to_clipboard(dm)

            # Click Message button to open the chat — user just Ctrl+V to paste
            msg_opened = _click_message_btn(page)
            if msg_opened:
                print("  Message chat opened. User can Ctrl+V to paste.")
            else:
                print("  Could not click Message button — user can open chat manually.")

        return jsonify({
            "dm":         dm,
            "lead":       lead,
            "clipboard":  clipboard,
            "msg_opened": msg_opened,
            "sent_today": _auto_state["sent_today"],
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/auto_skip", methods=["POST"])
def auto_skip():
    """Skip current lead (mark as Skip) and return next lead info without opening browser."""
    try:
        wb   = _auto_state["wb"]
        ws   = _auto_state["ws"]
        cols = _auto_state["cols"]
        row  = _auto_state["current_row"]

        if ws is None:
            return jsonify({"error": "No Excel loaded."}), 400

        if row is not None and cols["contacted"] is not None:
            ws.cell(row=row, column=cols["contacted"] + 1, value="Skip")
            _save_wb()

        _auto_state["current_row"] = None

        lead = _next_lead(ws, cols)
        if not lead:
            return jsonify({"error": "No more leads — all done!"}), 200

        _auto_state["current_row"] = lead["row"]

        total = sum(1 for r in ws.iter_rows(min_row=2) if str(r[cols["url"]].value or "").strip())
        sent  = sum(1 for r in ws.iter_rows(min_row=2)
                    if cols["contacted"] is not None and
                    str(r[cols["contacted"]].value or "").strip().lower() in ("yes","y","sent","1","true"))

        return jsonify({"ok": True, "lead": lead, "sent": sent, "total": total})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/auto_mark_sent", methods=["POST"])
def auto_mark_sent():
    """Mark current lead as sent in Excel."""
    try:
        wb   = _auto_state["wb"]
        ws   = _auto_state["ws"]
        cols = _auto_state["cols"]
        row  = _auto_state["current_row"]

        if ws is None or row is None:
            return jsonify({"ok": False, "error": "Nothing to mark."}), 400

        if cols["contacted"] is not None:
            ws.cell(row=row, column=cols["contacted"] + 1, value="Yes")
        if cols.get("sent_date") is not None:
            ws.cell(row=row, column=cols["sent_date"] + 1, value=_today_str())
        _save_wb()

        _auto_state["sent_today"] += 1
        _auto_state["current_row"] = None

        sent  = sum(1 for r in ws.iter_rows(min_row=2)
                    if cols["contacted"] is not None and
                    str(r[cols["contacted"]].value or "").strip().lower() in ("yes","y","sent","1","true"))
        total = sum(1 for r in ws.iter_rows(min_row=2) if str(r[cols["url"]].value or "").strip())

        return jsonify({"ok": True, "sent_today": _auto_state["sent_today"], "sent": sent, "total": total})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Owner Emails section ──────────────────────────────────────────────────────
# A separate, self-contained flow for leads whose contact is an email address.
# It scans the lead's Company Facebook page for a personal line, writes one short
# owner email, and copies it to the clipboard. It uses its own _email_state so it
# never disturbs the Auto Sender (DM) flow.

def _save_email_wb():
    """Save the Owner-Emails workbook, retrying briefly if the file is locked."""
    import time as _t
    wb   = _email_state.get("wb")
    path = _email_state.get("excel_path")
    if wb is None or not path:
        return
    for _ in range(6):
        try:
            wb.save(path); return
        except PermissionError:
            _t.sleep(0.4)
    raise PermissionError(
        "Can't write to the Excel file — it's open in Excel (or still syncing in "
        "OneDrive). Close the spreadsheet, wait for OneDrive, then click again.")


def _email_counts(ws, cols):
    total = sum(
        1 for r in ws.iter_rows(min_row=2)
        if (cols.get("contact") is not None and "@" in str(r[cols["contact"]].value or ""))
        or (cols.get("contact_type") is not None and "email" in str(r[cols["contact_type"]].value or "").lower()))
    sent = sum(
        1 for r in ws.iter_rows(min_row=2)
        if cols["contacted"] is not None and
        str(r[cols["contacted"]].value or "").strip().lower() in ("yes", "y", "sent", "1", "true"))
    return total, sent


@app.route("/email_load", methods=["POST"])
def email_load():
    """Load Excel for the Owner Emails section; return the first email lead."""
    try:
        data = request.get_json() or {}
        path = (data.get("path") or "").strip()
        if not path:
            for f in _Path(__file__).parent.glob("*.xlsx"):
                if any(k in f.name.lower() for k in ("lead", "email", "contact", "prospect", "outreach")):
                    path = str(f); break
        if not path or not _Path(path).exists():
            return jsonify({"error": f"Excel file not found: {path}"}), 400
        wb, ws, cols = _load_excel(path)
        _email_state.update({"excel_path": path, "wb": wb, "ws": ws, "cols": cols, "current_row": None})
        lead = _next_email_lead(ws, cols)
        if not lead:
            return jsonify({"error": "No email leads found in this file."}), 200
        total, sent = _email_counts(ws, cols)
        return jsonify({"lead": lead, "total": total, "sent": sent})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/email_next", methods=["POST"])
def email_next():
    """Scan the next email lead's Company FB page and write the owner email."""
    try:
        ws   = _email_state["ws"]
        cols = _email_state["cols"]
        if ws is None:
            return jsonify({"error": "Load your Excel file first."}), 400

        prev_row = _email_state["current_row"]
        if prev_row is not None and cols["contacted"] is not None:
            ws.cell(row=prev_row, column=cols["contacted"] + 1, value="Skip")
            _save_email_wb()

        lead = _next_email_lead(ws, cols, skip_row=prev_row)
        if not lead:
            return jsonify({"error": "No more email leads — all done!"}), 200
        _email_state["current_row"] = lead["row"]

        scan_url = (lead.get("url") or "").strip()

        # No Company FB page → write from spreadsheet data only, no Chrome.
        if not scan_url.startswith("http"):
            subject, body = _build_email("", "", lead)
            clipboard = copy_to_clipboard(body)
            return jsonify({"subject": subject, "email": lead.get("email", ""),
                            "body": body, "lead": lead, "clipboard": clipboard,
                            "sent_today": _email_state["sent_today"]})

        from playwright.sync_api import sync_playwright
        import socket
        connected = False
        for _ in range(5):
            try:
                s = socket.create_connection(("127.0.0.1", CDP_PORT), timeout=2); s.close()
                connected = True; break
            except:
                import time as _t2; _t2.sleep(1)
        if not connected:
            return jsonify({"error": "Chrome not found. Double-click START UK.bat, wait for Chrome to open, then try again."}), 500

        with sync_playwright() as p:
            browser = p.chromium.connect_over_cdp(f"http://localhost:{CDP_PORT}")
            ctx     = browser.contexts[0]
            for _pg in list(ctx.pages):
                _u = _pg.url.lower()
                if ("facebook.com/" in _u and "messenger.com" not in _u and
                        "/messages" not in _u and "localhost" not in _u):
                    try: _pg.close()
                    except Exception: pass
            page = ctx.new_page()
            try:
                page.goto(scan_url, wait_until="domcontentloaded", timeout=45000)
            except Exception:
                try:
                    page.goto(scan_url, wait_until="commit", timeout=30000)
                except Exception as e:
                    return jsonify({"error": f"Could not load page: {e}"}), 500
            page.wait_for_timeout(1000)
            try: page.keyboard.press("Escape")
            except Exception: pass
            try:
                page.evaluate(r"""() => {
                    if (!document.getElementById('dm-overlay-killer')) {
                        const st = document.createElement('style'); st.id = 'dm-overlay-killer';
                        st.textContent = '[role="dialog"]{display:none!important;}html,body{overflow:auto!important;}';
                        (document.head || document.documentElement).appendChild(st);
                    }
                    const nuke = () => document.querySelectorAll('[role="dialog"]').forEach(el => el.remove());
                    nuke(); const obs = new MutationObserver(nuke);
                    obs.observe(document.documentElement, { childList: true, subtree: true });
                    setTimeout(() => obs.disconnect(), 6000);
                }""")
            except Exception: pass
            try:
                text = page.evaluate("""() => {
                    ['script','style','noscript','svg','iframe'].forEach(t => document.querySelectorAll(t).forEach(el => el.remove()));
                    return document.body ? document.body.innerText : '';
                }""")
                import re as _re
                text = _re.sub(r"\n{3,}", "\n\n", text).strip()[:12000]
            except:
                text = ""

            text_data = text
            if lead["niche"]: text_data = f"--- TRADE TYPE ---\n{lead['niche']}\n\n" + text_data
            if lead["city"]:  text_data += f"\n\n--- LISTED CITY ---\n{lead['city']}"

            opener = ""
            try:
                import time as _tw
                page.evaluate("window.scrollTo(0, 850)")
                _tw.sleep(1.4)
                shot = page.screenshot(type="jpeg", quality=72)
                page.evaluate("window.scrollTo(0, 0)")
                raw = call_gemini(text_data=text_data, images=[(shot, "image/jpeg")])
                opener = patch_fallback_opener(raw.rstrip(".!,") + ".", text_data, lead.get("city", ""))
            except Exception as _se:
                print(f"  Email opener skipped: {_se}")

            try:
                page_title = page.evaluate(
                    "() => { const h = document.querySelector('h1');"
                    " const o = document.querySelector('meta[property=\"og:title\"]');"
                    " return (h && h.innerText) || (o && o.content) || document.title || ''; }")
            except Exception:
                page_title = ""

            subject, body = _build_email(opener, text, lead, page_title)
            clipboard = copy_to_clipboard(body)

        return jsonify({"subject": subject, "email": lead.get("email", ""),
                        "body": body, "lead": lead, "clipboard": clipboard,
                        "sent_today": _email_state["sent_today"]})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/email_skip", methods=["POST"])
def email_skip():
    """Skip the current email lead and return the next one (no Chrome)."""
    try:
        ws = _email_state["ws"]; cols = _email_state["cols"]; row = _email_state["current_row"]
        if ws is None:
            return jsonify({"error": "No Excel loaded."}), 400
        if row is not None and cols["contacted"] is not None:
            ws.cell(row=row, column=cols["contacted"] + 1, value="Skip"); _save_email_wb()
        _email_state["current_row"] = None
        lead = _next_email_lead(ws, cols)
        if not lead:
            return jsonify({"error": "No more email leads — all done!"}), 200
        _email_state["current_row"] = lead["row"]
        total, sent = _email_counts(ws, cols)
        return jsonify({"ok": True, "lead": lead, "sent": sent, "total": total})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/email_mark_sent", methods=["POST"])
def email_mark_sent():
    """Mark the current email lead as sent."""
    try:
        ws = _email_state["ws"]; cols = _email_state["cols"]; row = _email_state["current_row"]
        if ws is None or row is None:
            return jsonify({"ok": False, "error": "Nothing to mark."}), 400
        if cols["contacted"] is not None:
            ws.cell(row=row, column=cols["contacted"] + 1, value="Yes")
        if cols.get("sent_date") is not None:
            ws.cell(row=row, column=cols["sent_date"] + 1, value=_today_str())
        _save_email_wb()
        _email_state["sent_today"] += 1; _email_state["current_row"] = None
        total, sent = _email_counts(ws, cols)
        return jsonify({"ok": True, "sent_today": _email_state["sent_today"], "sent": sent, "total": total})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Follow-up tracking ───────────────────────────────────────────────────────
# A lead is "due" for a nudge when it's been contacted, hasn't replied, the
# sequence isn't finished (max 3 nudges), and enough days have passed since the
# LAST touch. Gaps widen: nudge #1 at 3 days, #2 at +7, #3 at +11.
_FOLLOWUP_GAP = {0: 3, 1: 7, 2: 11}

@app.route("/followups_due", methods=["POST"])
def followups_due():
    """Return every contacted-but-no-reply lead that's due for a follow-up."""
    try:
        ws   = _auto_state["ws"]
        cols = _auto_state["cols"]
        if ws is None:
            return jsonify({"error": "Load your leads on the Auto Sender tab first."}), 400

        from datetime import date
        today = date.today()
        due = []
        for row in ws.iter_rows(min_row=2):
            r   = row[0].row
            url = str(row[cols["url"]].value or "").strip()
            if not url.startswith("http"):
                continue
            contacted = (str(row[cols["contacted"]].value or "").strip().lower()
                         if cols["contacted"] is not None else "")
            if contacted not in ("yes", "y", "sent", "1", "true"):
                continue  # only leads we've actually messaged
            replied = (str(row[cols["replied"]].value or "").strip().lower()
                       if cols.get("replied") is not None else "")
            if replied in ("yes", "y", "1", "true"):
                continue  # they replied — never nudge
            fups = _to_int(row[cols["followups"]].value) if cols.get("followups") is not None else 0
            if fups >= FOLLOWUP_TOUCHES:
                continue  # sequence finished
            last = _parse_date(row[cols["sent_date"]].value) if cols.get("sent_date") is not None else None
            if last is None:
                continue  # no date logged — can't schedule
            days = (today - last).days
            if days < _FOLLOWUP_GAP.get(fups, 99):
                continue  # not due yet
            due.append({
                "row":   r,
                "name":  str(row[cols["name"]].value or "").strip(),
                "url":   url,
                "city":  str(row[cols["city"]].value or "").strip() if cols["city"] is not None else "",
                "niche": str(row[cols["niche"]].value or "").strip() if cols["niche"] is not None else "",
                "days":  days,
                "touch": fups + 1,
            })
        due.sort(key=lambda d: -d["days"])  # most overdue first
        return jsonify({"due": due, "count": len(due)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/followup_send", methods=["POST"])
def followup_send():
    """Prepare the next nudge for one lead: copy it to the clipboard, log the
    touch (advance the cadence), and return the text + profile URL to open."""
    try:
        ws   = _auto_state["ws"]
        cols = _auto_state["cols"]
        if ws is None:
            return jsonify({"error": "Load your leads first."}), 400
        r = int(request.get_json(force=True).get("row", 0))
        if r < 2:
            return jsonify({"error": "Bad row."}), 400

        name = str(ws.cell(row=r, column=cols["name"] + 1).value or "").strip()
        url  = str(ws.cell(row=r, column=cols["url"] + 1).value or "").strip()
        niche = (str(ws.cell(row=r, column=cols["niche"] + 1).value or "").strip()
                 if cols.get("niche") is not None else "")
        fups = _to_int(ws.cell(row=r, column=cols["followups"] + 1).value) if cols.get("followups") is not None else 0
        text = _followup_text(fups + 1, name, niche)

        if cols.get("followups") is not None:
            ws.cell(row=r, column=cols["followups"] + 1, value=fups + 1)
        if cols.get("sent_date") is not None:
            ws.cell(row=r, column=cols["sent_date"] + 1, value=_today_str())
        _save_wb()

        clipboard = copy_to_clipboard(text)
        return jsonify({"ok": True, "text": text, "url": url, "clipboard": clipboard, "touch": fups + 1})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/mark_replied", methods=["POST"])
def mark_replied():
    """Mark a lead as replied so it drops off the follow-up list for good."""
    try:
        ws   = _auto_state["ws"]
        cols = _auto_state["cols"]
        if ws is None:
            return jsonify({"error": "Load your leads first."}), 400
        r = int(request.get_json(force=True).get("row", 0))
        if r < 2 or cols.get("replied") is None:
            return jsonify({"error": "Can't mark this row."}), 400
        ws.cell(row=r, column=cols["replied"] + 1, value="Yes")
        _save_wb()
        return jsonify({"ok": True})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=False, port=5001)

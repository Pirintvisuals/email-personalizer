"""
Landscaper Email Campaign Generator — Stateless Flask App
Works on Vercel (serverless) and locally.
Frontend drives the loop: one /process-record call per landscaper.
"""

import io
import os
import re
import json
from datetime import datetime

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Flask, render_template, request, jsonify, send_file
from google import genai

# ─── CONFIG ────────────────────────────────────────────────────────────────────

# Load .env if present (local dev)
_env = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.isfile(_env):
    with open(_env) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _v = _line.split("=", 1)
                os.environ.setdefault(_k.strip(), _v.strip())

GEMINI_MODEL    = "gemini-2.5-flash"
REQUEST_TIMEOUT = 10

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# ─── FLASK APP ─────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

# Lazy Gemini client — initialised on first use so a missing env var
# doesn't crash the entire serverless function at import time.
_gemini_client = None

def get_gemini_client():
    global _gemini_client
    if _gemini_client is None:
        api_key = os.environ.get("GEMINI_API_KEY", "")
        if not api_key:
            raise RuntimeError("GEMINI_API_KEY environment variable is not set.")
        _gemini_client = genai.Client(api_key=api_key)
    return _gemini_client


# ─── HELPERS ───────────────────────────────────────────────────────────────────

def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def fetch_website(url: str) -> dict:
    result = {
        "success": False, "url": url, "title": "", "body_text": "",
        "has_contact_form": False, "has_booking": False,
        "has_cta": False, "page_meta": "", "error": "", "email": "",
    }
    if not url or url.strip() in ("", "N/A", "n/a", "-"):
        result["error"] = "No URL provided"
        return result
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    result["url"] = url

    try:
        resp = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        resp.raise_for_status()
    except requests.exceptions.SSLError:
        try:
            resp = requests.get(url.replace("https://", "http://"),
                                headers=HEADERS, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            resp.raise_for_status()
        except Exception as e:
            result["error"] = f"SSL + HTTP fallback failed: {str(e)[:120]}"
            return result
    except Exception as e:
        result["error"] = str(e)[:150]
        return result

    try:
        soup = BeautifulSoup(resp.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "head",
                         "noscript", "iframe", "svg"]):
            tag.decompose()
        title_tag = soup.find("title")
        result["title"] = clean_text(title_tag.get_text()) if title_tag else ""
        meta = soup.find("meta", attrs={"name": re.compile(r"description", re.I)})
        if meta and meta.get("content"):
            result["page_meta"] = clean_text(meta["content"])
        body = soup.find("body")
        raw = body.get_text(separator=" ") if body else soup.get_text(separator=" ")
        result["body_text"] = clean_text(raw)[:3000]
        result["has_contact_form"] = bool(soup.find_all("form"))
        page_lower = resp.text.lower()
        result["has_booking"] = any(k in page_lower for k in
            ["book", "schedule", "appointment", "calendar", "reserve",
             "get a quote", "free quote", "request a quote", "estimate"])
        result["has_cta"] = any(k in page_lower for k in
            ["call us", "contact us", "get started", "free estimate",
             "call now", "get a quote", "request service"])
        # Extract email addresses from page
        raw_emails = re.findall(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', resp.text)
        skip_email = ['noreply', 'no-reply', 'donotreply', 'example.com', 'w3.org', 'schema.org', 'sentry', 'wix.com', 'wordpress']
        clean_emails = [e for e in raw_emails if not any(s in e.lower() for s in skip_email)]
        result["email"] = clean_emails[0] if clean_emails else ""
        result["success"] = True
    except Exception as e:
        result["error"] = f"Parse error: {str(e)[:120]}"
    return result


def build_prompt(company: str, contact: str, site_data: dict, town: str = "", stars: str = "", review_count: str = "") -> str:
    if site_data["success"]:
        site_summary = f"""WEBSITE DATA:
- Title: {site_data['title']}
- Meta description: {site_data['page_meta']}
- Has contact/lead form: {site_data['has_contact_form']}
- Has booking / quote CTA: {site_data['has_booking']}
- Has call-to-action: {site_data['has_cta']}
- Page text excerpt:
{site_data['body_text'][:2500]}"""
    else:
        site_summary = f"WEBSITE FETCH FAILED: {site_data['error']}"

    # Clean up review count — spreadsheet stores as negative e.g. -26 → 26
    try:
        review_count_clean = str(abs(int(float(review_count)))) if review_count else ""
    except (ValueError, TypeError):
        review_count_clean = review_count

    stars_clean = stars.strip() if stars else ""

    town_line   = f"Town/Area: {town}" if town else ""
    stars_line  = f"Star Rating: {stars_clean}" if stars_clean else ""
    review_line = f"Review Count: {review_count_clean}" if review_count_clean else ""
    extra = "\n".join(filter(None, [town_line, stars_line, review_line]))

    first_name = contact.split()[0] if contact and contact.strip() and contact != "there" else "{FIRST_NAME}"

    # Pick an opener style based on company name hash so it's consistent but varied
    style = hash(company) % 4

    area = town or "the area"

    if review_count_clean and stars_clean:
        stats = f"{review_count_clean} reviews and {stars_clean} stars"
        if style == 0:
            opener_hint = (
                f'Pattern interrupt — lead with a bold honest observation. E.g. '
                f'"{review_count_clean} reviews, {stars_clean} stars — '
                f'honestly that\'s a solid rep for a landscaper in {area}. '
                f'My bet is you\'re still taking calls from people who were never going to book."'
            )
        elif style == 1:
            opener_hint = (
                f'Ask a direct question that makes them think. E.g. '
                f'"Quick one — {company} has {stats} on Google. '
                f'How many of the enquiries you got last month actually turned into paid work?"'
            )
        elif style == 2:
            opener_hint = (
                f'Lead with empathy and a sharp insight. E.g. '
                f'"Ran across {company} online — {stats} in {area}. '
                f'I know how it goes at your level: plenty of people calling, '
                f'half of them just price-shopping or tyre-kicking."'
            )
        else:
            opener_hint = (
                f'Start with a genuine compliment then flip it. E.g. '
                f'"{stats} — {company} clearly knows what they\'re doing in {area}. '
                f'Still, I\'d guess at least a few hours a week go on enquiries that go nowhere."'
            )
    elif stars_clean:
        if style % 2 == 0:
            opener_hint = (
                f'Pattern interrupt with their star rating. E.g. '
                f'"{stars_clean} stars on Google — {company} is clearly doing good work in {area}. '
                f'The question is whether every enquiry you\'re getting is actually worth your time."'
            )
        else:
            opener_hint = (
                f'Direct question using their rating. E.g. '
                f'"Looked up {company} — {stars_clean} stars. Solid. '
                f'How much of your week goes on leads that never actually convert?"'
            )
    elif review_count_clean:
        opener_hint = (
            f'Reference their review count with a sharp hook. E.g. '
            f'"{review_count_clean} Google reviews — clearly a trusted name. '
            f'I\'d still bet some of those enquiries weren\'t worth picking up the phone for."'
        )
    else:
        opener_hint = (
            'Write a pattern-interrupting opener based on something specific from their website — '
            'their services, a specific project type, or something genuine. '
            'End with a question or bold observation about lead quality. '
            'DO NOT use "I hope", "reaching out", or generic openers.'
        )

    # Pick subject line style based on hash
    subj_style = hash(company + "subj") % 8
    if subj_style == 0:
        subject_instruction = f'Ultra-simple, 1-3 words — e.g. "thoughts on this?" or "quick one" or "{first_name}?" (use their first name if known)'
    elif subj_style == 1:
        subject_instruction = f'Personalized value — e.g. "better leads for {company}?" or "tyre-kickers vs real clients"'
    elif subj_style == 2:
        subject_instruction = f'Curiosity — e.g. "noticed your reviews" or "working with a landscaper in {area}"'
    elif subj_style == 3:
        subject_instruction = f'Pattern interrupt — reply style, e.g. "re: {company}" (looks like a follow-up thread)'
    elif subj_style == 4:
        subject_instruction = f'First name only if known — e.g. "{first_name}" or leave completely blank for highest open rate'
    elif subj_style == 5:
        subject_instruction = f'Company reference — e.g. "{company}" or "{company} — quick question"'
    elif subj_style == 6:
        subject_instruction = f'Stats-driven — e.g. "your {review_count_clean or ""} reviews" or "5-star leads" or "300% better leads — {company}"'
    else:
        subject_instruction = f'Conversion-focused — e.g. "quality over quantity?" or "{first_name} — conversion rate" or "tyre-kickers?"'

    return f"""You are a direct, no-fluff B2B cold email writer for a lead-filtering service helping landscaping companies stop wasting time on tyre-kickers.

Company: {company}
Contact name (use ONLY if clearly a real first name, otherwise leave blank): {contact}
Website: {site_data['url']}
{extra}

{site_summary}

YOUR TASK — return a JSON object with exactly THREE keys:

1. "contact_name" — If you can clearly identify the owner or main contact's FIRST NAME from the website content, return just the first name (e.g. "John"). If not confident, return "".

2. "subject" — Subject line ONLY (no "Subject:" prefix). Style to use: {subject_instruction}

3. "email_body" — The email body only (NO subject line in here). Follow this structure exactly:

Hey [first name if known, otherwise just skip the greeting entirely],

[PERSONALISED OPENER — {opener_hint}]

Just reaching out because we help landscapers stop wasting time on tyre-kickers, so they only ever speak to homeowners who are serious about getting work done.

We just started working with All Things Outside Ltd over in Ilfracombe — they were getting plenty of enquiries but most weren't going anywhere. We're already tracking them towards 300% better quality leads.

Can I send a quick video explaining how it works?

Milan
+447478075473

Rules:
- Use the contact's first name if known — NEVER write "there" or leave a blank placeholder
- If no name known, skip the greeting line entirely and start with the opener
- Opener: 1–2 sentences, sharp and specific to THIS company
- Everything after the opener is FIXED word for word as above
- Total 80–110 words in email_body
- NO "I hope", NO "just following up", NO fluff, NO "Hi,"

Return ONLY valid JSON. No markdown fences. No extra text."""


def generate_emails(company: str, contact: str, site_data: dict, town: str = "", stars: str = "", review_count: str = "") -> dict:
    prompt = build_prompt(company, contact, site_data, town, stars, review_count)
    raw = ""
    try:
        response = get_gemini_client().models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
        )
        raw = response.text.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw.strip())
        return json.loads(raw)
    except json.JSONDecodeError:
        try:
            match = re.search(r'\{.*\}', raw, re.DOTALL)
            if match:
                return json.loads(match.group())
        except Exception:
            pass
        return {"subject": "ERROR", "email_body": f"JSON parse failed. Raw: {raw[:200]}"}
    except Exception as e:
        return {"subject": "ERROR", "email_body": str(e)[:200]}


def build_excel_bytes(records: list) -> bytes:
    """Generate the campaign Excel in memory and return raw bytes."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Send-ready (the 3 columns you actually need for sending) ──
    ws = wb.active
    ws.title = "Send Ready"
    send_cols = ["Email Address", "Subject", "Email Body"]
    header_fill  = PatternFill("solid", fgColor="1B5E20")
    header_font  = Font(color="FFFFFF", bold=True, size=11)
    for ci, col in enumerate(send_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fill_e = PatternFill("solid", fgColor="E8F5E9")
    fill_o = PatternFill("solid", fgColor="FFFFFF")
    for ri, rec in enumerate(records, 2):
        if rec.get("skipped") or not rec.get("Email", "").strip():
            continue
        fill = fill_e if ri % 2 == 0 else fill_o
        for ci, val in enumerate([
            rec.get("Email", ""),
            rec.get("subject", ""),
            rec.get("email_body", rec.get("email_1", "")),
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for ci, w in enumerate([30, 40, 90], 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    ws.row_dimensions[1].height = 30
    for ri in range(2, len(records) + 2):
        ws.row_dimensions[ri].height = 130
    ws.freeze_panes = "A2"

    # ── Sheet 2: Full data (all fields for reference) ──
    ws2 = wb.create_sheet("Full Data")
    all_cols = [
        "Company Name", "Contact Name", "Email", "Website URL",
        "Phone Number", "Town", "Stars", "Review Count", "Subject", "Email Body"
    ]
    hf2 = PatternFill("solid", fgColor="1F4E79")
    for ci, col in enumerate(all_cols, 1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.fill = hf2
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ri, rec in enumerate(records, 2):
        fill = fill_e if ri % 2 == 0 else fill_o
        for ci, val in enumerate([
            rec.get("Company Name", ""),  rec.get("Contact Name", ""),
            rec.get("Email", ""),         rec.get("Website URL", ""),
            rec.get("Phone Number", ""),  rec.get("Town", ""),
            rec.get("Stars", ""),         rec.get("Review Count", ""),
            rec.get("subject", ""),
            rec.get("email_body", rec.get("email_1", "")),
        ], 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for ci, w in enumerate([28, 20, 28, 35, 18, 16, 8, 10, 35, 80], 1):
        ws2.column_dimensions[ws2.cell(row=1, column=ci).column_letter].width = w
    ws2.row_dimensions[1].height = 30
    for ri in range(2, len(records) + 2):
        ws2.row_dimensions[ri].height = 120
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── ROUTES ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    """Read Excel in memory, return headers + all rows as JSON for frontend mapping."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Please upload an Excel file (.xlsx or .xls)"}), 400

    try:
        raw = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return jsonify({"error": f"Could not read file: {str(e)}"}), 400

    if not all_rows:
        return jsonify({"error": "File is empty"}), 400

    # Skip leading blank rows before the header
    while all_rows and all(c is None or str(c).strip() == "" for c in all_rows[0]):
        all_rows = all_rows[1:]

    if not all_rows:
        return jsonify({"error": "File is empty"}), 400

    headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]

    # If all headers are blank the sheet has no header row — generate names and use all rows as data
    if all(h == "" for h in headers):
        num_cols = len(headers)
        headers = [f"Col {i+1}" for i in range(num_cols)]
        data_start = 0  # first row IS data
    else:
        data_start = 1  # first row was the header

    sample = [str(c).strip() if c is not None else "" for c in (all_rows[data_start] if len(all_rows) > data_start else [])]

    def clean_val(v):
        """Strip leading bullet/separator chars (·, -, •) and whitespace."""
        return re.sub(r'^[\s·•\-–—]+', '', str(v).strip()).strip()

    # Collect cleaned values per column across all data rows (up to 10 rows)
    num_cols = len(headers)
    col_samples = [[] for _ in range(num_cols)]
    for row in all_rows[data_start:data_start+10]:
        for ci, cell in enumerate(row):
            if ci < num_cols:
                v = clean_val(cell) if cell is not None else ""
                if v:
                    col_samples[ci].append(v)

    # Auto-guess column indices from header names first
    guesses = {}
    for i, h in enumerate(headers):
        hl = h.lower()
        if any(k in hl for k in ["company", "business"]) and "contact" not in hl:
            guesses.setdefault("company", i)
        elif any(k in hl for k in ["contact", "first name", "person"]):
            guesses.setdefault("contact", i)
        elif any(k in hl for k in ["website", "url", "web", "site"]):
            guesses.setdefault("website", i)
        elif any(k in hl for k in ["phone", "tel", "mobile", "cell"]):
            guesses.setdefault("phone", i)
        elif any(k in hl for k in ["email", "e-mail", "mail"]):
            guesses.setdefault("email", i)
        elif any(k in hl for k in ["town", "city", "area", "location", "region", "address"]):
            guesses.setdefault("town", i)
        elif any(k in hl for k in ["star"]) and "review" not in hl:
            guesses.setdefault("stars", i)
        elif any(k in hl for k in ["review", "rating", "score", "count"]):
            guesses.setdefault("review_count", i)

    # Scan all collected column values to fill any gaps
    for i, vals in enumerate(col_samples):
        if not vals:
            continue
        # Count how many values in this column match each pattern
        url_hits   = sum(1 for v in vals if re.search(r'https?://', v) or v.startswith("www."))
        phone_hits = sum(1 for v in vals if re.search(r'\+?\d[\d\s\-]{7,}', v))
        email_hits = sum(1 for v in vals if re.search(r'@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', v))
        num_hits   = sum(1 for v in vals if re.match(r'^-?\d+(\.\d+)?$', v))
        # Short alphabetic words = likely town names
        town_hits  = sum(1 for v in vals
                         if re.match(r'^[A-Za-z][A-Za-z\s\-]{1,25}$', v.split('·')[0].strip())
                         and not re.search(r'\d', v.split('·')[0]))

        if url_hits >= len(vals) // 2:
            guesses.setdefault("website", i)
        elif email_hits >= max(1, len(vals) // 2):
            guesses.setdefault("email", i)
        elif phone_hits >= len(vals) // 2:
            guesses.setdefault("phone", i)
        elif i == 0 and vals:
            guesses.setdefault("company", i)
        elif town_hits >= len(vals) // 2 and "town" not in guesses:
            guesses.setdefault("town", i)
        elif num_hits >= len(vals) // 2:
            h = headers[i].lower() if i < len(headers) else ""
            if any(k in h for k in ["star"]) and "review" not in h:
                guesses.setdefault("stars", i)
            elif any(k in h for k in ["review", "rating", "score", "count"]):
                guesses.setdefault("review_count", i)

    # Return all data rows as plain arrays (frontend applies col mapping)
    data_rows = []
    for row in all_rows[data_start:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        data_rows.append([str(c).strip() if c is not None else "" for c in row])

    # Build a sample array using the first non-empty value per column
    best_sample = [next((v for v in col_samples[i] if v), "") for i in range(num_cols)]

    return jsonify({
        "headers":  headers,
        "sample":   best_sample,
        "guesses":  guesses,
        "rows":     data_rows,
        "total":    len(data_rows),
    })


@app.route("/process-record", methods=["POST"])
def process_record():
    """Process a single landscaper: fetch website + generate emails.
    Called once per record from the frontend loop — stays well under Vercel timeout."""
    def strip_bullets(v):
        return re.sub(r'^[\s·•\-–—]+', '', str(v or "")).strip()

    data         = request.get_json()
    company      = strip_bullets(data.get("company", "")) or "Unknown Company"
    contact      = strip_bullets(data.get("contact", "")) or "there"
    website      = strip_bullets(data.get("website", ""))
    phone        = strip_bullets(data.get("phone", ""))
    town_raw     = strip_bullets(data.get("town", "")).split("·")[0].strip()
    # Discard values that are business age ("5+ years in business") not a town name
    town         = "" if re.search(r'\d+\+?\s*years', town_raw, re.IGNORECASE) else town_raw
    stars        = strip_bullets(data.get("stars", ""))
    review_count = strip_bullets(data.get("review_count", ""))
    email        = strip_bullets(data.get("email", ""))

    site_data = fetch_website(website)
    # If no email from spreadsheet, try to scrape it from the website
    if not email:
        email = site_data.get("email", "")

    # Skip generation entirely if no email address found
    if not email:
        return jsonify({
            "Company Name":  company,
            "Contact Name":  contact,
            "Email":         "",
            "Website URL":   website,
            "Phone Number":  phone,
            "Town":          town,
            "Stars":         stars,
            "Review Count":  review_count,
            "site_ok":       site_data["success"],
            "site_error":    site_data.get("error", ""),
            "email_1":       "SKIPPED — no email address found",
            "skipped":       True,
        })

    result = generate_emails(company, contact, site_data, town, stars, review_count)

    # Use AI-extracted contact name if we didn't have one from the spreadsheet
    final_contact = contact
    if (not final_contact or final_contact == "there"):
        ai_name = result.get("contact_name", "").strip()
        if ai_name and len(ai_name.split()) <= 2 and ai_name.replace(" ", "").isalpha():
            final_contact = ai_name

    return jsonify({
        "Company Name":  company,
        "Contact Name":  final_contact,
        "Email":         email,
        "Website URL":   website,
        "Phone Number":  phone,
        "Town":          town,
        "Stars":         stars,
        "Review Count":  review_count,
        "site_ok":       site_data["success"],
        "site_error":    site_data.get("error", ""),
        "subject":       result.get("subject", ""),
        "email_body":    result.get("email_body", result.get("email_1", "")),
    })


@app.route("/export", methods=["POST"])
def export_excel():
    """Receive all completed records, return Excel file as download."""
    records = request.get_json()
    if not records:
        return jsonify({"error": "No records to export"}), 400

    try:
        xlsx_bytes = build_excel_bytes(records)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        io.BytesIO(xlsx_bytes),
        as_attachment=True,
        download_name=f"campaign_{ts}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─── ENTRY POINT ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    host = "0.0.0.0" if os.environ.get("PORT") else "127.0.0.1"
    print("\n" + "=" * 55)
    print("  Landscaper Email Campaign Generator")
    print(f"  Running at http://{host}:{port}")
    print("=" * 55 + "\n")
    app.run(debug=False, host=host, port=port)
